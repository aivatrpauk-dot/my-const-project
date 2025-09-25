#PNL Ошибка
import logging
import asyncio
import openpyxl
import pandas as pd
from openpyxl import load_workbook
import io
from datetime import datetime, timedelta, datetime as dt
from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, InputFile
from tg_bot.models import sessionmaker, engine
from tg_bot.models import Shop, Advertisement, Penalty
from tg_bot.models import (
    TaxSystemType, TaxSystemSetting,
    ProductCost, RegularExpense, OneTimeExpense
)
from tg_bot.keyboards.pnl_menu import pnl_period_keyboard
from tg_bot.services.wb_api import fetch_full_report, fetch_report_detail_by_period
from tg_bot.states.pnl_states import PNLStates
from dateutil.relativedelta import relativedelta
from tg_bot.models import Order

logger = logging.getLogger(__name__)

# Главное меню P&L
async def pnl_callback(callback: types.CallbackQuery, state: FSMContext):
    # Проверяем выбран ли магазин
    async with state.proxy() as data:
        if 'shop' not in data:
            await callback.answer("❌ Сначала выберите магазин", show_alert=True)
            return
    
    keyboard = pnl_period_keyboard()
    await callback.message.answer(
        "📊 <b>Расчёт прибыли и убытков (P&L)</b>\n\n"
        "Выберите период для расчета:",
        reply_markup=keyboard
    )
    await PNLStates.waiting_for_period.set()

async def generate_pnl_excel_report(shop_id: int, shop_api_token: str, start_date: datetime, end_date: datetime, shop_name: str, full_data=None):
    """Генерация Excel-отчета PNL"""

    session = sessionmaker()(bind=engine)
    try:
        
        # Если full_data не передан, получаем его из API
        if full_data is None:
            logger.info("Получаем данные из API...")
            full_data = await fetch_full_report(shop_api_token, start_date, end_date, shop_id)
            if not full_data:
                logger.error("Не удалось получить данные из API")
                return None
    
        report_data = full_data['finance']
        orders = full_data['orders']
        sales = full_data['sales']
        
        # Отладочная информация о входящих данных
        logger.info(f"Входящие данные:")
        logger.info(f"  - Finance записей: {len(report_data)}")
        logger.info(f"  - Orders записей: {len(orders)}")
        logger.info(f"  - Sales записей: {len(sales)}")
        
        if report_data:
            sample_finance = report_data[0]
            logger.info(f"  - Пример finance записи: {sample_finance}")
        
        if orders:
            sample_order = orders[0]
            logger.info(f"  - Пример order записи: {sample_order}")
            
        if sales:
            sample_sale = sales[0]
            logger.info(f"  - Пример sale записи: {sample_sale}")
        
        # Определяем период
        DATE_FROM = start_date.strftime("%Y-%m-%d")
        DATE_TO = end_date.strftime("%Y-%m-%d")

        df_orders = pd.DataFrame(orders)
        df_sales = pd.DataFrame(sales)
        df_fin = pd.DataFrame(report_data)
        
        # Отладочная информация о DataFrame
        logger.info(f"DataFrame информация:")
        logger.info(f"  - df_orders: {len(df_orders)} строк, колонки: {list(df_orders.columns) if not df_orders.empty else 'пустой'}")
        logger.info(f"  - df_sales: {len(df_sales)} строк, колонки: {list(df_sales.columns) if not df_sales.empty else 'пустой'}")
        logger.info(f"  - df_fin: {len(df_fin)} строк, колонки: {list(df_fin.columns) if not df_fin.empty else 'пустой'}")

        def determine_period_type(date_from, date_to):
            """Определяет тип периода и возвращает расширенный период для загрузки."""
            start_dt = dt.strptime(date_from, "%Y-%m-%d")
            end_dt = dt.strptime(date_to, "%Y-%m-%d")
            
            days_diff = (end_dt - start_dt).days
            
            if days_diff == 0:
                # День - загружаем неделю
                period_start = start_dt - timedelta(days=6)
                period_end = start_dt + timedelta(days=6)
                period_type = "день"
            elif days_diff <= 7:
                # Неделя - загружаем месяц
                period_start = start_dt - timedelta(days=30)
                period_end = end_dt + timedelta(days=30)
                period_type = "неделя"
            elif days_diff <= 31:
                # Месяц - загружаем квартал
                period_start = start_dt - timedelta(days=90)
                period_end = end_dt + timedelta(days=90)
                period_type = "месяц"
            else:
                # Год - загружаем год
                period_start = start_dt - timedelta(days=365)
                period_end = end_dt + timedelta(days=365)
                period_type = "год"
            
            return period_start.strftime("%Y-%m-%d"), period_end.strftime("%Y-%m-%d"), period_type

        # Определяем расширенный период
        extended_date_from, extended_date_to, period_type = determine_period_type(DATE_FROM, DATE_TO)

        # day‑столбец
        if not df_orders.empty:
            df_orders["day"] = pd.to_datetime(df_orders["date"]).dt.date
        else:
            df_orders["day"] = pd.to_datetime("2025-01-01").date()
            
        if not df_sales.empty:
            df_sales["day"] = pd.to_datetime(df_sales["date"]).dt.date
        else:
            df_sales["day"] = pd.to_datetime("2025-01-01").date()
        
        # Проверяем наличие поля rr_dt в финансовых данных
        if not df_fin.empty:
            if "rr_dt" in df_fin.columns:
                df_fin["day"] = pd.to_datetime(df_fin["rr_dt"]).dt.date
            elif "date" in df_fin.columns:
                df_fin["day"] = pd.to_datetime(df_fin["date"]).dt.date
            else:
                # Если нет поля даты, создаем пустой столбец day
                df_fin["day"] = pd.to_datetime("2025-01-01").date()
        else:
            df_fin["day"] = pd.to_datetime("2025-01-01").date()

        # 4.1 Orders
        if not df_orders.empty:
            grp_orders = (
                df_orders.groupby("day")
                        .agg(
                            order_sum = ("priceWithDisc", "sum"),
                            ordered_qty = ("srid", "size"),
                            orders_cnt = ("srid", "nunique")
                        )
            )
        else:
            # Создаем пустой DataFrame с нужными столбцами
            grp_orders = pd.DataFrame(columns=['order_sum', 'ordered_qty', 'orders_cnt'])

        # 4.2 Sales (только продажи, не возвраты)
        if not df_sales.empty:
            sales_mask = df_sales["saleID"].str.startswith("S", na=False)
            grp_sales = (
                df_sales[sales_mask]
                .groupby("day")
                .agg(
                    sales_sum = ("priceWithDisc", "sum"),
                    sold_qty = ("saleID", "size"),
                    for_pay = ("forPay", "sum")
                )
            )
            
            # Также получаем возвраты из sales данных
            returns_mask = df_sales["saleID"].str.startswith("R", na=False)
            if returns_mask.any():
                df_returns_sales = df_sales[returns_mask]
                df_returns_sales["day"] = pd.to_datetime(df_returns_sales["date"]).dt.date
                grp_returns_sales = (
                    df_returns_sales.groupby("day")
                    .agg(
                        returns_sum_sales = ("priceWithDisc", "sum"),
                        returns_qty_sales = ("saleID", "size")
                    )
                )
            else:
                grp_returns_sales = pd.DataFrame(columns=['returns_sum_sales', 'returns_qty_sales'])
        else:
            # Создаем пустой DataFrame с нужными столбцами
            grp_sales = pd.DataFrame(columns=['sales_sum', 'sold_qty', 'for_pay'])
            grp_returns_sales = pd.DataFrame(columns=['returns_sum_sales', 'returns_qty_sales'])

        # 4.3 Finance — с правильным расчетом возвратов, удержаний и штрафов
        if not df_fin.empty:
            # Фильтруем данные для текущего периода
            current_start = start_date.date()
            current_end = end_date.date()
            
            # Рассчитываем возвраты, удержания и штрафы из finance данных
            returns_data = []
            deductions_data = []
            penalties_data = []
            
            for item in report_data:
                if not isinstance(item, dict):
                    continue
                
                # Проверяем дату записи
                item_date = None
                if "rr_dt" in item:
                    item_date = datetime.strptime(item["rr_dt"][:10], "%Y-%m-%d").date()
                elif "date" in item:
                    item_date = datetime.strptime(item["date"][:10], "%Y-%m-%d").date()
                
                if item_date and current_start <= item_date <= current_end:
                    doc_type = item.get("doc_type_name", "").lower()
                    
                    # Возвраты
                    if "возврат" in doc_type or "return" in doc_type:
                        returns_data.append({
                            "day": item_date,
                            "returns_sum": item.get("retail_price_withdisc_rub", 0) * item.get("quantity", 0),
                            "returns_qty": item.get("quantity", 0)
                        })
                    
                    # Прочие удержания (deduction)
                    if item.get("deduction", 0) != 0:
                        deductions_data.append({
                            "day": item_date,
                            "deduction": item.get("deduction", 0)
                        })
                    
                    # Штрафы (penalty + другие штрафы)
                    penalty_amount = item.get("penalty", 0)
                    if penalty_amount != 0:
                        penalties_data.append({
                            "day": item_date,
                            "penalty": penalty_amount
                        })
            
            # Создаем DataFrame для возвратов, удержаний и штрафов
            df_returns = pd.DataFrame(returns_data)
            df_deductions = pd.DataFrame(deductions_data)
            df_penalties = pd.DataFrame(penalties_data)
            
            # Группируем по дням
            if not df_returns.empty:
                grp_returns = df_returns.groupby("day").agg({
                    "returns_sum": "sum",
                    "returns_qty": "sum"
                })
            else:
                grp_returns = pd.DataFrame(columns=['returns_sum', 'returns_qty'])
            
            if not df_deductions.empty:
                grp_deductions = df_deductions.groupby("day").agg({
                    "deduction": "sum"
                }).rename(columns={"deduction": "additional_deductions"})
            else:
                grp_deductions = pd.DataFrame(columns=['additional_deductions'])
            
            if not df_penalties.empty:
                grp_penalties = df_penalties.groupby("day").agg({
                    "penalty": "sum"
                }).rename(columns={"penalty": "additional_penalties"})
            else:
                grp_penalties = pd.DataFrame(columns=['additional_penalties'])
            
            # Основная группировка finance данных
            grp_fin = (
                df_fin.groupby("day")
                    .agg(
                        delivery_cost = ("delivery_rub", "sum"),
                        storage = ("storage_fee", "sum"),
                        penalty = ("penalty", "sum"),
                        acceptance = ("acceptance", "sum"),
                        pay_for_goods = ("ppvz_for_pay", "sum"),
                        deduction = ("deduction", "sum")  # Добавляем столбец deduction
                    )
                    .apply(pd.to_numeric, errors="coerce")
            )
        else:
            # Создаем пустые DataFrame с нужными столбцами
            grp_fin = pd.DataFrame(columns=['delivery_cost', 'storage', 'penalty', 'acceptance', 'pay_for_goods', 'deduction'])
            grp_returns = pd.DataFrame(columns=['returns_sum', 'returns_qty'])
            grp_deductions = pd.DataFrame(columns=['additional_deductions'])
            grp_penalties = pd.DataFrame(columns=['additional_penalties'])

        # ────────────────────────────────────────────────────────────────────
        # 5. Объединяем и рассчитываем
        # ────────────────────────────────────────────────────────────────────
        all_daily = (
            grp_orders
            .join(grp_sales, how="outer")
            .join(grp_fin, how="outer")
            .join(grp_returns, how="outer")
            .join(grp_returns_sales, how="outer")
            .join(grp_deductions, how="outer")
            .join(grp_penalties, how="outer")
            .fillna(0)
            .infer_objects(copy=False)
        )
        
        # Объединяем возвраты из finance и sales
        all_daily["returns_sum"] = all_daily["returns_sum"] + all_daily["returns_sum_sales"]
        all_daily["returns_qty"] = all_daily["returns_qty"] + all_daily["returns_qty_sales"]

        all_daily["buyout_pct"] = (
            all_daily["sold_qty"] / all_daily["ordered_qty"]
        ).fillna(0)

        all_daily["total_to_pay"] = (
            all_daily["pay_for_goods"]
            - all_daily[["delivery_cost", "storage", "penalty", "acceptance"]].sum(axis=1)
        )

        # Фильтруем данные для текущего периода
        current_start = start_date.date()
        current_end = end_date.date()
        daily = all_daily[(all_daily.index >= current_start) & (all_daily.index <= current_end)]
        
        # Отладочная информация
        logger.info(f"Период: {current_start} - {current_end}")
        logger.info(f"Всего записей в all_daily: {len(all_daily)}")
        logger.info(f"Записей в daily после фильтрации: {len(daily)}")
        if not daily.empty:
            logger.info(f"Индексы daily: {daily.index.tolist()}")
        else:
            logger.info(f"Индексы all_daily: {all_daily.index.tolist() if not all_daily.empty else 'Пустой'}")
            logger.info(f"Проблема: нет данных в all_daily для периода {current_start} - {current_end}")

        # Получаем рекламу и штрафы из базы данных
        advert = sum(
            i.amount for i in session.query(Advertisement)
            .filter(Advertisement.shop_id == shop_id)
            .filter(Advertisement.date >= start_date)
            .filter(Advertisement.date <= end_date)
        )

        stops = sum(
            i.amount for i in session.query(Penalty)
            .filter(Penalty.shop_id == shop_id)
            .filter(Penalty.date >= start_date.date())
            .filter(Penalty.date <= end_date.date())
        )

        # Получаем настройки налоговой системы
        tax_setting = session.query(TaxSystemSetting).filter(TaxSystemSetting.shop_id == shop_id).first()
        
        # Рассчитываем налог
        revenue = daily["sales_sum"].sum()
        if tax_setting:
            if tax_setting.tax_system == TaxSystemType.USN_6:
                tax_rate = 0.06
                print("tax_rate = ", tax_rate)
            elif tax_setting.tax_system == TaxSystemType.NO_TAX:
                tax_rate = 0.0
                print("tax_rate = ", tax_rate)
            elif tax_setting.tax_system == TaxSystemType.CUSTOM:
                tax_rate = tax_setting.custom_percent / 100
                print("tax_rate = ", tax_rate)
            else:
                tax_rate = 0.0
                print("пошло в else")
        else:
            tax_rate = 0.0
            
        tax = revenue * tax_rate
        
        # Себестоимость (как в analytics.py)
        cost_of_goods = 0
        articles = {}
        
        # Собираем артикулы для расчета себестоимости
        for item in report_data:
            if not isinstance(item, dict):
                continue
            article = item.get("nm_id")
            quantity = item.get("quantity", 0)
            if article and quantity:
                if article not in articles:
                    articles[article] = 0
                articles[article] += quantity

        # Рассчитываем себестоимость
        for article, quantity in articles.items():
            try:
                supp_article = session.query(Order).filter(Order.nmId == int(article)).first().supplierArticle
                product_cost = (
                    session.query(ProductCost)
                    .filter(ProductCost.shop_id == shop_id, ProductCost.article == supp_article)
                    .first()
                )
                if product_cost:
                    cost_of_goods += product_cost.cost * quantity
            except:
                pass
        
        # Формируем метрики как в Test.py
        monthly_data = {
            "Заказы": daily["order_sum"].sum(),
            "Выкупы": daily["sales_sum"].sum(),
            "Комиссия": daily["pay_for_goods"].sum(),
            "Себестоймость": cost_of_goods,  # Теперь с правильным расчетом
            "Налог": tax,
            "Логистика": daily["delivery_cost"].sum(),
            "Хранение": daily["storage"].sum(),
            "Штрафы и корректировки": daily["penalty"].sum() + stops,
            "Реклама": advert,
            "К перечислению": daily["sales_sum"].sum() - daily["pay_for_goods"].sum() - daily["delivery_cost"].sum() - daily["storage"].sum() - (daily["penalty"].sum() + stops) - advert - cost_of_goods,
            "Чистая прибыль": daily["sales_sum"].sum() - daily["pay_for_goods"].sum() - daily[["delivery_cost", "storage", "penalty"]].sum().sum() - advert - tax - cost_of_goods
        }

        # Получаем данные для сравнения (предыдущий период)
        if period_type == "день":
            # Для дня - предыдущий день
            prev_start = current_start - timedelta(days=1)
            prev_end = prev_start
            period_days = 1
        elif period_type == "неделя":
            # Для недели - предыдущая неделя
            prev_start = current_start - timedelta(days=7)
            prev_end = current_start - timedelta(days=1)
            period_days = 7
        elif period_type == "месяц":
            # Для месяца - предыдущий месяц
            prev_start = current_start - timedelta(days=30)
            prev_end = current_start - timedelta(days=1)
            period_days = 30
        else:  # год
            # Для года - предыдущий год
            prev_start = current_start - timedelta(days=365)
            prev_end = current_start - timedelta(days=1)
            period_days = 365

        # Получаем данные за предыдущий период
        prev_daily = all_daily[(all_daily.index >= prev_start) & (all_daily.index <= prev_end)]

        # Рассчитываем налог для предыдущего периода
        prev_revenue = prev_daily["sales_sum"].sum()
        prev_tax = prev_revenue * tax_rate

        # Получаем рекламу за предыдущий период
        prev_advert = sum(
            i.amount for i in session.query(Advertisement)
            .filter(Advertisement.shop_id == shop_id)
            .filter(Advertisement.date >= (start_date - timedelta(days=period_days)).date())
            .filter(Advertisement.date < start_date.date())
            .all()
        )

        # Себестоимость для предыдущего периода (как в analytics.py)
        prev_cost_of_goods = 0
        prev_articles = {}
        
        # Собираем артикулы для расчета себестоимости за предыдущий период
        for item in report_data:
            if not isinstance(item, dict):
                continue
            article = item.get("nm_id")
            quantity = item.get("quantity", 0)
            # Проверяем, что товар был продан в предыдущем периоде
            item_date = datetime.strptime(item.get("sale_dt", "2025-01-01")[:10], "%Y-%m-%d").date()
            if prev_start <= item_date <= prev_end and article and quantity:
                if article not in prev_articles:
                    prev_articles[article] = 0
                prev_articles[article] += quantity

        # Рассчитываем себестоимость для предыдущего периода
        for article, quantity in prev_articles.items():
            try:
                supp_article = session.query(Order).filter(Order.nmId == int(article)).first().supplierArticle
                product_cost = (
                    session.query(ProductCost)
                    .filter(ProductCost.shop_id == shop_id, ProductCost.article == supp_article)
                    .first()
                )
                if product_cost:
                    prev_cost_of_goods += product_cost.cost * quantity
            except:
                pass

        # Формируем метрики за предыдущий период
        prev_monthly_data = {
            "Заказы": prev_daily["order_sum"].sum(),
            "Выкупы": prev_daily["sales_sum"].sum(),
            "Комиссия": prev_daily["pay_for_goods"].sum(),
            "Себестоймость": prev_cost_of_goods,  # Теперь с правильным расчетом
            "Налог": prev_tax,
            "Логистика": prev_daily["delivery_cost"].sum(),
            "Хранение": prev_daily["storage"].sum(),
            "Штрафы и корректировки": prev_daily["penalty"].sum(),
            "Реклама": prev_advert,
            "К перечислению": prev_daily["sales_sum"].sum() - prev_daily["pay_for_goods"].sum() - prev_daily["delivery_cost"].sum() - prev_daily["storage"].sum() - prev_daily["penalty"].sum() - prev_advert - prev_cost_of_goods,
            "Чистая прибыль": prev_daily["sales_sum"].sum() - prev_daily["pay_for_goods"].sum() - prev_daily[["delivery_cost", "storage", "penalty"]].sum().sum() - prev_tax - prev_advert - prev_cost_of_goods
        }

        # Рассчитываем относительные изменения
        def calculate_relative_change(current, previous):
            if previous == 0:
                return 0
            return (current - previous) / previous

        relative_changes = {}
        for metric in monthly_data.keys():
            relative_changes[metric] = calculate_relative_change(
                monthly_data[metric], 
                prev_monthly_data[metric]
            )

        # Проверяем наличие данных только для логирования
        if not report_data:
            logger.warning("Нет данных из API, но продолжаем обработку")

        # Загружаем шаблон
        try:
            wb = load_workbook("pnl_template.xlsx")
        except FileNotFoundError:
            logger.error("Файл шаблона pnl_template.xlsx не найден")
            return None

        ws = wb.active

        # Функция для расчета себестоимости за конкретный день
        def calculate_cost_for_day(day_date, report_data):
            day_cost = 0
            day_articles = {}
            
            for item in report_data:
                if not isinstance(item, dict):
                    continue
                article = item.get("nm_id")
                quantity = item.get("quantity", 0)
                item_date = datetime.strptime(item.get("sale_dt", "2025-01-01")[:10], "%Y-%m-%d").date()
                
                # Убеждаемся, что day_date тоже является date объектом
                if isinstance(day_date, datetime):
                    day_date = day_date.date()
                
                if item_date == day_date and article and quantity:
                    if article not in day_articles:
                        day_articles[article] = 0
                    day_articles[article] += quantity
            
            for article, quantity in day_articles.items():
                try:
                    supp_article = session.query(Order).filter(Order.nmId == int(article)).first().supplierArticle
                    product_cost = (
                        session.query(ProductCost)
                        .filter(ProductCost.shop_id == shop_id, ProductCost.article == supp_article)
                        .first()
                    )
                    if product_cost:
                        day_cost += product_cost.cost * quantity
                except:
                    pass
            
            return day_cost

        # Функция для расчета себестоимости за конкретный месяц
        def calculate_cost_for_month(month_num, year, report_data):
            month_cost = 0
            month_articles = {}
            
            month_start = datetime(year, month_num, 1)
            if month_num == 12:
                month_end = datetime(year + 1, 1, 1) - timedelta(days=1)
            else:
                month_end = datetime(year, month_num + 1, 1) - timedelta(days=1)
            
            for item in report_data:
                if not isinstance(item, dict):
                    continue
                article = item.get("nm_id")
                quantity = item.get("quantity", 0)
                item_date = datetime.strptime(item.get("sale_dt", "2025-01-01")[:10], "%Y-%m-%d")
                
                if month_start <= item_date <= month_end and article and quantity:
                    if article not in month_articles:
                        month_articles[article] = 0
                    month_articles[article] += quantity
            
            for article, quantity in month_articles.items():
                try:
                    supp_article = session.query(Order).filter(Order.nmId == int(article)).first().supplierArticle
                    product_cost = (
                        session.query(ProductCost)
                        .filter(ProductCost.shop_id == shop_id, ProductCost.article == supp_article)
                        .first()
                    )
                    if product_cost:
                        month_cost += product_cost.cost * quantity
                except:
                    pass
            
            return month_cost

        # Заполняем A1 - период
        period_text = f"{start_date.strftime('%d.%m.%Y')}-{end_date.strftime('%d.%m.%Y')}"
        ws["A1"] = period_text

        # Рассчитываем метрики за весь период
        total_orders = daily["order_sum"].sum()
        total_sales = daily["sales_sum"].sum()
        total_cost = sum(calculate_cost_for_day(date, report_data) for date in daily.index)
        # Правильный расчет комиссии WB
        # Комиссия = ppvz_sales_commission + ppvz_vw + ppvz_vw_nds
        total_commission = sum(
            item.get("ppvz_sales_commission", 0) +
            item.get("ppvz_vw", 0) +
            item.get("ppvz_vw_nds", 0)
            for item in report_data
        )
        # Правильный расчет возвратов
        total_returns = daily["returns_sum"].sum()
        
        # Также добавляем возвраты из sales данных за текущий период
        current_sales_returns = [sale for sale in sales if 
                               sale.get("saleID", "").startswith("R") and
                               start_date <= datetime.strptime(sale.get("date", "2025-01-01")[:10], "%Y-%m-%d") <= end_date]
        for sale in current_sales_returns:
            total_returns += sale.get("priceWithDisc", 0)
        total_advert = sum(
            i.amount for i in session.query(Advertisement)
            .filter(Advertisement.shop_id == shop_id)
            .filter(Advertisement.date >= start_date.date())
            .filter(Advertisement.date <= end_date.date())
            .all()
        )
        total_logistics = daily["delivery_cost"].sum()
        total_storage = daily["storage"].sum()
        # Правильный расчет прочих удержаний
        total_deductions = daily["deduction"].sum() + daily["additional_deductions"].sum()
        # Правильный расчет штрафов (включая штрафы из БД)
        total_penalties = daily["penalty"].sum() + daily["additional_penalties"].sum() + stops
        total_payout = total_sales - total_commission - total_logistics - total_storage - total_penalties - total_advert - total_cost
        total_tax = total_sales * tax_rate
        total_profit = total_sales - total_commission - total_logistics - total_storage - total_penalties - total_tax - total_advert - total_cost

        # Заполняем строку 2 (метрики за весь период)
        ws["B2"] = total_orders
        ws["C2"] = total_sales
        ws["D2"] = total_cost
        ws["E2"] = total_commission
        ws["F2"] = total_returns
        ws["G2"] = total_advert
        ws["H2"] = total_logistics
        ws["I2"] = total_storage
        ws["J2"] = total_deductions
        ws["K2"] = total_penalties
        ws["L2"] = total_payout
        ws["M2"] = total_tax
        ws["N2"] = total_profit

        # Рассчитываем разницу с прошлым периодом
        # Данные за прошлый период уже включены в full_data благодаря расширению периода в wb_api.py
        period_days = (end_date - start_date).days + 1
        prev_start_date = start_date - timedelta(days=period_days)
        prev_end_date = start_date - timedelta(days=1)
        
        # Фильтруем данные за прошлый период из уже полученных данных
        def parse_date_safe(date_str, default='2025-01-01'):
            """Безопасный парсинг даты с обработкой различных форматов"""
            if not date_str:
                return datetime.strptime(default, '%Y-%m-%d')
            try:
                # Убираем время если есть (T21:48:39)
                date_only = date_str.split('T')[0] if 'T' in date_str else date_str
                return datetime.strptime(date_only, '%Y-%m-%d')
            except ValueError:
                return datetime.strptime(default, '%Y-%m-%d')
        
        prev_report_data = [item for item in report_data if 
                           prev_start_date <= parse_date_safe(item.get('rr_dt', item.get('date'))) <= prev_end_date]
        prev_orders = [order for order in orders if 
                      prev_start_date <= parse_date_safe(order.get('date')) <= prev_end_date]
        prev_sales = [sale for sale in sales if 
                     prev_start_date <= parse_date_safe(sale.get('date')) <= prev_end_date]
        
        # Обрабатываем данные за прошлый период
        prev_df_orders = pd.DataFrame(prev_orders)
        prev_df_sales = pd.DataFrame(prev_sales)
        prev_df_fin = pd.DataFrame(prev_report_data)
        
        # Рассчитываем метрики за прошлый период
        prev_total_orders = prev_df_orders["priceWithDisc"].sum() if not prev_df_orders.empty else 0
        prev_total_sales = prev_df_sales["priceWithDisc"].sum() if not prev_df_sales.empty else 0
        prev_total_cost = sum(calculate_cost_for_day(date, prev_report_data) for date in pd.date_range(prev_start_date, prev_end_date))
        # Правильный расчет комиссии WB для прошлого периода
        prev_total_commission = sum(
            item.get("ppvz_sales_commission", 0) +
            item.get("ppvz_vw", 0) +
            item.get("ppvz_vw_nds", 0)
            for item in prev_report_data
        )
        # Правильный расчет возвратов для прошлого периода
        prev_total_returns = 0
        for item in prev_report_data:
            if not isinstance(item, dict):
                continue
            doc_type = item.get("doc_type_name", "").lower()
            if "возврат" in doc_type or "return" in doc_type:
                prev_total_returns += item.get("retail_price_withdisc_rub", 0) * item.get("quantity", 0)
        
        # Также добавляем возвраты из sales данных за предыдущий период
        prev_sales_returns = [sale for sale in prev_sales if sale.get("saleID", "").startswith("R")]
        for sale in prev_sales_returns:
            prev_total_returns += sale.get("priceWithDisc", 0)
        
        prev_total_advert = sum(
            i.amount for i in session.query(Advertisement)
            .filter(Advertisement.shop_id == shop_id)
            .filter(Advertisement.date >= prev_start_date)
            .filter(Advertisement.date <= prev_end_date)
            .all()
        )
        prev_total_logistics = prev_df_fin["delivery_rub"].sum() if not prev_df_fin.empty else 0
        prev_total_storage = prev_df_fin["storage_fee"].sum() if not prev_df_fin.empty else 0
        # Правильный расчет штрафов для прошлого периода
        prev_total_penalties = prev_df_fin["penalty"].sum() if not prev_df_fin.empty else 0
        # Правильный расчет прочих удержаний для прошлого периода
        prev_total_deductions = sum(item.get("deduction", 0) for item in prev_report_data)
        
        prev_total_tax = prev_total_sales * tax_rate
        prev_total_profit = prev_total_sales - prev_total_commission - prev_total_logistics - prev_total_storage - prev_total_penalties - prev_total_tax - prev_total_advert - prev_total_cost

        # Заполняем строку 3 (разница с прошлым периодом) - сохраняем исходное форматирование
        values = [
            ("B4", total_orders - prev_total_orders),
            ("C4", total_sales - prev_total_sales),
            ("D4", total_cost - prev_total_cost),
            ("E4", total_commission - prev_total_commission),
            ("F4", total_returns - prev_total_returns),
            ("G4", total_advert - prev_total_advert),
            ("H4", total_logistics - prev_total_logistics),
            ("I4", total_storage - prev_total_storage),
            ("J4", total_deductions - prev_total_deductions),
            ("K4", total_penalties - prev_total_penalties),
            ("L4", total_payout - (prev_total_sales - prev_total_commission - prev_total_logistics - prev_total_storage - prev_total_penalties - prev_total_advert - prev_total_cost)),
            ("M4", total_tax - prev_total_tax),
            ("N4", total_profit - prev_total_profit)
        ]
        
        for cell_ref, value in values:
            cell = ws[cell_ref]
            original_format = cell.number_format  # Сохраняем исходное форматирование
            cell.value = value
            cell.number_format = original_format  # Восстанавливаем исходное форматирование

        # Создаем полный диапазон дат для заполнения всех дней периода
        date_range = pd.date_range(start=current_start, end=current_end, freq='D')
        
        # Логируем информацию о диапазоне дат
        logger.info(f"Создан диапазон дат: {len(date_range)} дней с {current_start} по {current_end}")
        logger.info(f"Дни в диапазоне: {[d.strftime('%Y-%m-%d') for d in date_range]}")
        
        # Заполняем данные по дням начиная с 5 строки
        row = 5
        for date in date_range:
            # Конвертируем pandas.Timestamp в datetime.date для сравнения
            date_date = date.date()
            
            # Получаем данные за день (если есть)
            if date_date in daily.index:
                day_data = daily.loc[date_date]
                logger.info(f"День {date.strftime('%Y-%m-%d')}: найдены данные")
            else:
                # Если данных нет, создаем пустые значения
                day_data = pd.Series({
                    'order_sum': 0,
                    'sales_sum': 0,
                    'delivery_cost': 0,
                    'storage': 0,
                    'penalty': 0,
                    'returns_sum': 0,
                    'deduction': 0,
                    'additional_penalties': 0,
                    'additional_deductions': 0
                })
                logger.info(f"День {date.strftime('%Y-%m-%d')}: данных нет, заполняем нулями")
            
            day_cost = calculate_cost_for_day(date_date, report_data)
            
            # Получаем рекламу за день
            day_advert = sum(
                i.amount for i in session.query(Advertisement)
                .filter(Advertisement.shop_id == shop_id)
                .filter(Advertisement.date == date_date)
                .all()
            )
            
            # Рассчитываем метрики за день
            day_orders = day_data.get("order_sum", 0)
            day_sales = day_data.get("sales_sum", 0)
            
            # Правильный расчет комиссии WB для дня
            day_report_data = [item for item in report_data if 
                              parse_date_safe(item.get('rr_dt', item.get('date'))).date() == date_date]
            day_commission = sum(
                item.get("ppvz_sales_commission", 0) +
                item.get("ppvz_vw", 0) +
                item.get("ppvz_vw_nds", 0)
                for item in day_report_data
            )
            
            day_logistics = day_data.get("delivery_cost", 0)
            day_storage = day_data.get("storage", 0)
            day_penalties = day_data.get("penalty", 0)
            day_returns = day_data.get("returns_sum", 0)
            
            # Также добавляем возвраты из sales данных за этот день
            day_sales_returns = [sale for sale in sales if 
                               sale.get("saleID", "").startswith("R") and
                               datetime.strptime(sale.get("date", "2025-01-01")[:10], "%Y-%m-%d").date() == date_date]
            for sale in day_sales_returns:
                day_returns += sale.get("priceWithDisc", 0)
            day_deductions = day_data.get("deduction", 0)
            
            # Объединяем штрафы из основного finance и дополнительных штрафов
            additional_penalties = day_data.get("additional_penalties", 0)
            additional_deductions = day_data.get("additional_deductions", 0)
            total_day_penalties = day_penalties + additional_penalties
            total_day_deductions = day_deductions + additional_deductions
            day_tax = day_sales * tax_rate
            day_profit = day_sales - day_commission - day_logistics - day_storage - total_day_penalties - day_tax - day_advert - day_cost
            
            # Заполняем строку
            ws[f"A{row}"] = date.strftime("%d.%m.%Y")  # Дата в формате ДД.ММ.ГГГГ
            ws[f"B{row}"] = day_orders
            ws[f"C{row}"] = day_sales
            ws[f"D{row}"] = day_cost
            ws[f"E{row}"] = day_commission
            ws[f"F{row}"] = day_returns
            ws[f"G{row}"] = day_advert
            ws[f"H{row}"] = day_logistics
            ws[f"I{row}"] = day_storage
            ws[f"J{row}"] = total_day_deductions
            ws[f"K{row}"] = total_day_penalties
            ws[f"L{row}"] = day_sales - day_commission - day_logistics - day_storage - total_day_penalties - day_advert - day_cost
            ws[f"M{row}"] = day_tax
            ws[f"N{row}"] = day_profit
            
            # Копируем стиль с 5-й строки (шаблон) для всех ячеек
            if row > 5:
                # Копируем форматирование ячеек
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
                    source_cell = ws[f"{col}5"]
                    target_cell = ws[f"{col}{row}"]
                    
                    # Копируем форматирование безопасно
                    try:
                        if source_cell.font:
                            target_cell.font = source_cell.font.copy()
                        if source_cell.border:
                            target_cell.border = source_cell.border.copy()
                        if source_cell.fill:
                            target_cell.fill = source_cell.fill.copy()
                        if source_cell.number_format:
                            target_cell.number_format = source_cell.number_format
                        if source_cell.alignment:
                            target_cell.alignment = source_cell.alignment.copy()
                    except Exception as style_error:
                        # Если не удается скопировать стиль, используем базовое форматирование
                        logger.warning(f"Не удалось скопировать стиль для ячейки {col}{row}: {style_error}")
                        # Устанавливаем базовое форматирование
                        target_cell.number_format = '#,##0.00'
            
            row += 1

        logger.info(f"Заполнено строк: {row - 5} (с 5-й по {row-1}-ю строку)")

        return wb
        
    except Exception as e:
        logger.error(f"Ошибка генерации PNL отчета: {e}")
        return None
    finally:
        session.close()

# Обработка выбора периода
async def select_pnl_period_callback(callback: types.CallbackQuery, state: FSMContext):
    period_type = callback.data.split('_')[1]  # day, week, month, year
    await callback.message.edit_text(text="Подождите около 10 секунд, пока произведем подсчёт данных... (иногда дольше, но не более 2х минут)")
    
    # Определяем периоды
    now = datetime.utcnow()
    if period_type == "week":
        # Находим последний понедельник
        start_week = now - timedelta(days=now.isoweekday() - 1)
        start_date = datetime(start_week.year, start_week.month, start_week.day)
        end_date = now
        period_name = "неделю"
    elif period_type == "month":
        start_date = datetime(now.year, now.month, 1)
        end_date = now
        period_name = "месяц"
    elif period_type == "year":
        start_date = datetime(now.year, 1, 1)
        end_date = now
        period_name = "год"
    else:  # day
        start_date = now - timedelta(days=1)
        end_date = now
        period_name = "день"
    
    async with state.proxy() as data:
        shop_id = data['shop']['id']
        shop_name = data['shop']['name'] or f"Магазин {shop_id}"
        shop_api_token = data['shop']['api_token']
    
    # Показываем сообщение о загрузке
    await callback.message.edit_text(
        f"📊 <b>Генерация PNL отчета</b>\n\n"
        f"Магазин: {shop_name}\n"
        f"Период: за {period_name}\n\n"
        "Подождите, идет сбор и обработка данных..."
    )
    
    # Генерируем Excel отчет
    wb = await generate_pnl_excel_report(shop_id, shop_api_token, start_date, end_date, shop_name)
    
    if not wb:
        await callback.message.edit_text(
            "❌ <b>Не удалось сгенерировать отчет</b>\n\n"
            "Возможные причины:\n"
            "1. Нет данных за выбранный период\n"
            "2. Проблемы с подключением к базе данных\n"
            "3. Файл шаблона pnl_template.xlsx не найден"
        )
        return
    
    # Сохраняем в буфер
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    # Формируем имя файла
    safe_shop_name = "".join(c for c in shop_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
    filename = f"PNL.xlsx"
    
    # Отправляем файл
    file = InputFile(file_stream, filename=filename)
    await callback.message.answer_document(
        file,
    )
    
    # Возвращаемся к меню PNL
    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("🔙 Назад", callback_data="pnl"))
    await callback.message.answer("✅ Отчет готов!", reply_markup=keyboard)

def register_pnl_handlers(dp):
    dp.register_callback_query_handler(pnl_callback, text="pnl", state="*")
    dp.register_callback_query_handler(
        select_pnl_period_callback, 
        lambda c: c.data.startswith("pnlperiod_"), 
        state=PNLStates.waiting_for_period
    )