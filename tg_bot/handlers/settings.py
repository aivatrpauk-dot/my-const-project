import logging
import io
import openpyxl
from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, InputFile
from tg_bot.models import Shop, sessionmaker, engine, User, Order
from tg_bot.models import (
    TaxSystemSetting, TaxSystemType,
    ProductCost,
    RegularExpense, RegularExpenseFrequency,
    OneTimeExpense
)
from tg_bot.states.settings_states import SettingsStates
from tg_bot.keyboards.settings_menu import (
    tax_system_keyboard,
    regular_expense_frequency_keyboard,
    settings_menu_keyboard
)
from datetime import datetime

TAX_MAPPING = {
    "usn_6": TaxSystemType.USN_6,
    "notax": TaxSystemType.NO_TAX
}

logger = logging.getLogger(__name__)
async def daily_reports_callback(callback: types.CallbackQuery, state):
    # Здесь логика включения/выключения ежедневных отчётов
    session = sessionmaker(bind=engine)()
    try:
        user = session.query(User).filter(User.telegram_id == callback.from_user.id).first()
        if user:
            # Переключаем состояние ежедневных отчетов
            user.daily_reports_enabled = not user.daily_reports_enabled
            session.commit()
            status = "включены" if user.daily_reports_enabled else "выключены"
            await callback.answer(f"Ежедневные отчёты {status}!", show_alert=True)
            
            # Дополнительная информация для пользователя
            if user.daily_reports_enabled:
                await callback.message.answer(
                    "📊 Автоматические отчеты включены!\n\n"
                    "📅 Еженедельные отчеты будут приходить каждую среду в 12:00\n"
                    "📅 Ежемесячные отчеты будут приходить каждое 3 число в 12:00\n\n"
                    "Отчеты будут содержать данные за прошедший период в формате Excel."
                )
            else:
                await callback.message.answer(
                    "❌ Автоматические отчеты отключены.\n"
                    "Вы больше не будете получать еженедельные и ежемесячные отчеты."
                )
            
            # Обновляем меню настроек с новым статусом
            try:
                await callback.message.edit_reply_markup(
                    reply_markup=settings_menu_keyboard(user.shops[0].id, user.daily_reports_enabled)
                )
            except:
                pass  # Игнорируем ошибки при обновлении клавиатуры
        else:
            await callback.answer("Пользователь не найден", show_alert=True)
    except Exception as e:
        logger.error(f"Ошибка при переключении ежедневных отчетов: {e}")
        await callback.answer("Произошла ошибка при изменении настроек", show_alert=True)
    finally:
        session.close()


async def settings_callback(callback: types.CallbackQuery, state: FSMContext):
    await SettingsStates.menu.set()
    Session = sessionmaker()
    session = Session(bind=engine)
    user = session.query(User).filter(User.telegram_id == callback.from_user.id).first()
    shop = session.query(Shop).filter(Shop.user_id==user.id).first()
    async with state.proxy() as data:
        data['shop'] = {
            'id': shop.id,
            'name': shop.name,
            'api_token': shop.api_token
        }
    session.close()
    try:
        await callback.message.edit_text(
            "⚙️ <b>Настройка параметров</b>\nВыберите опцию:",
            reply_markup=settings_menu_keyboard(shop.id, user.daily_reports_enabled)
        )
    except:
        await callback.message.delete()
        await callback.message.answer(
            "⚙️ <b>Настройка параметров</b>\nВыберите опцию:",
            reply_markup=settings_menu_keyboard(shop.id, user.daily_reports_enabled)
        )

async def back_to_settings(callback: types.CallbackQuery, state: FSMContext):
    await settings_callback(callback, state)

async def back_to_settings(callback: types.CallbackQuery, state: FSMContext):
    await SettingsStates.menu.set()
    Session = sessionmaker()
    session = Session(bind=engine)
    user = session.query(User).filter(User.telegram_id == callback.from_user.id).first()
    shop = session.query(Shop).filter(Shop.user_id==user.id).first()
    session.close()
    await callback.message.edit_text(
        "⚙️ <b>Настройка параметров</b>\nВыберите опцию:",
        reply_markup=settings_menu_keyboard(shop.id, user.daily_reports_enabled)
    )
async def back_to_settings_message(message: types.Message, state: FSMContext):
    await SettingsStates.menu.set()
    Session = sessionmaker()
    session = Session(bind=engine)
    user = session.query(User).filter(User.telegram_id == message.from_user.id).first()
    shop = session.query(Shop).filter(Shop.user_id==user.id).first()
    session.close()
    await message.answer(
        "⚙️ <b>Настройка параметров</b>\nВыберите опцию:",
        reply_markup=settings_menu_keyboard(shop.id, user.daily_reports_enabled)
    )

async def product_cost_callback_helper(message: types.Message):
    keyboard = InlineKeyboardMarkup(row_width=1)
    keyboard.add(
        InlineKeyboardButton("📤 Загрузить Excel", callback_data="upload_cost_excel"),
        InlineKeyboardButton("📝 Скачать шаблон", callback_data="download_cost_template"),
        InlineKeyboardButton("🔙 Назад", callback_data="back_to_settings")
    )
    await message.answer(
        "📦 <b>Управление себестоимостью артикулов</b>\n\n"
        "Выберите действие:",
        reply_markup=keyboard
    )

# Налоговая система
async def tax_settings_callback(callback: types.CallbackQuery, state: FSMContext):
    print(f"tax_settings_callback called with data: {callback.data}")
    session = sessionmaker()(bind=engine)
    try:
        async with state.proxy() as data:
            print(f"State data: {data}")
            if 'shop' not in data:
                await callback.answer("❌ Сначала выберите магазин", show_alert=True)
                return
            shop_id = data['shop']['id']
            print(f"Shop ID: {shop_id}")
        
        shop = session.query(Shop).filter(Shop.id == shop_id).first()
        if not shop:
            await callback.answer("❌ Магазин не найден", show_alert=True)
            return
        
        current_tax = shop.tax_settings.tax_system if shop.tax_settings else None
        print(f"Current tax: {current_tax}")
        
        text = "<b>Налоговая система</b>\n\n"
        if current_tax:
            text += f"Текущая система: <b>{current_tax.value}</b>\n\n"
        else:
            text += "❌ Налоговая система не выбрана\n\n"
        
        text += "Выберите систему налогообложения:"
        
        try:
            await callback.message.edit_text(text, reply_markup=tax_system_keyboard(current_tax))
        except Exception as edit_error:
            print(f"Edit error: {edit_error}")
            # Если не удалось отредактировать, отправляем новое сообщение
            try:
                await callback.message.delete()
            except:
                pass  # Игнорируем ошибку удаления
            await callback.message.answer(text, reply_markup=tax_system_keyboard(current_tax))
            return  # Выходим из функции, чтобы избежать повторного вызова
        
        await SettingsStates.tax_system.set()
    except Exception as e:
        print(f"Error in tax_settings_callback: {e}")
        await callback.answer(f"❌ Ошибка: {e}", show_alert=True)
    finally:
        session.close()

async def set_custom_tax_callback(callback: types.CallbackQuery, state: FSMContext):
    """Обработчик для ввода произвольного процента налога"""
    await callback.message.edit_text(
        "<b>Введите процент налога</b>\n\n"
        "Напишите число от 0 до 100 (например: 15 для 15%):",
        reply_markup=InlineKeyboardMarkup().add(
            InlineKeyboardButton("🔙 Назад", callback_data="tax_settings")
        )
    )
    await SettingsStates.waiting_for_tax_percent.set()

async def process_tax_percent(message: types.Message, state: FSMContext):
    """Обработчик ввода процента налога"""
    try:
        tax_percent = float(message.text.strip())
        print("Введенный custom_percent = ", tax_percent)  # ← ЗДЕСЬ
        
        if 0 <= tax_percent <= 100:
            # Сохраняем в состояние
            async with state.proxy() as data:
                data['custom_tax_percent'] = tax_percent
            
            # Сохраняем в базу данных
            session = sessionmaker()(bind=engine)
            try:
                shop_id = data['shop']['id']
                shop = session.query(Shop).filter(Shop.id == shop_id).first()
                
                if shop.tax_settings:
                    shop.tax_settings.tax_system = TaxSystemType.CUSTOM
                    shop.tax_settings.custom_percent = tax_percent
                else:
                    tax_setting = TaxSystemSetting(
                        shop_id=shop_id, 
                        tax_system=TaxSystemType.CUSTOM,
                        custom_percent=tax_percent
                    )
                    session.add(tax_setting)
                
                session.commit()
                await message.answer(f"✅ Налоговая ставка установлена: {tax_percent}%")
                await back_to_settings_message(message, state)
            finally:
                session.close()
        else:
            await message.answer("❌ Процент должен быть от 0 до 100. Попробуйте снова:")
    except ValueError:
        await message.answer("❌ Введите корректное число. Попробуйте снова:")

async def set_tax_system_callback(callback: types.CallbackQuery, state: FSMContext):
    tax_type_value = callback.data.split('_', 1)[1]
    
    # Сопоставляем значение с элементами перечисления
    tax_type = None
    if tax_type_value == "usn6":
        tax_type = TaxSystemType.USN_6
        print("tax_type = ", tax_type)
    elif tax_type_value == "notax":
        tax_type = TaxSystemType.NO_TAX
        print("tax_type = ", tax_type)
    elif tax_type_value == "custom":
        # Переходим к вводу процента
        await set_custom_tax_callback(callback, state)
        return        
    else:
        await callback.answer("❌ Неизвестный тип налоговой системы")
        return
    
    session = sessionmaker()(bind=engine)
    try:
        async with state.proxy() as data:
            shop_id = data['shop']['id']
        
        shop = session.query(Shop).filter(Shop.id == shop_id).first()
        if not shop:
            await callback.answer("❌ Магазин не найден", show_alert=True)
            return
        
        # Обновляем или создаем настройку
        if shop.tax_settings:
            shop.tax_settings.tax_system = tax_type
        else:
            tax_setting = TaxSystemSetting(shop_id=shop_id, tax_system=tax_type)
            session.add(tax_setting)
        
        session.commit()
        await callback.answer(f"✅ Налоговая система установлена: {tax_type.value}")
        await back_to_settings(callback, state)
    except Exception as e:
        logger.error(f"Ошибка установки налоговой системы: {e}")
        await callback.answer("❌ Произошла ошибка при установке системы")
    finally:
        session.close()
# Себестоимость артикулов
async def product_cost_callback(callback: types.CallbackQuery):
    keyboard = InlineKeyboardMarkup(row_width=1)
    keyboard.add(
        InlineKeyboardButton("📤 Загрузить Excel", callback_data="upload_cost_excel"),
        # InlineKeyboardButton("📥 Скачать текущие данные", callback_data="download_cost_excel"),
        InlineKeyboardButton("📝 Скачать шаблон", callback_data="download_cost_template"),
        InlineKeyboardButton("🔙 Назад", callback_data="back_to_settings")
    )
    await SettingsStates.product_cost.set()
    await callback.message.edit_text(
        " <b>Управление себестоимостью артикулов</b>\n\nСкачайте готовый шаблон для заполнния и нажмите загрузить Excel.",
        reply_markup=keyboard
    )

async def download_cost_template_callback(callback: types.CallbackQuery):
    # Создаем шаблон Excel
    session = sessionmaker(bind=engine)()
    shop = session.query(Shop).filter(Shop.user_id==session.query(User).filter(User.telegram_id==callback.from_user.id).first().id).first()
    articles = [str(item[0]) for item in session.query(Order.supplierArticle).filter(Order.shop_id==shop.id).distinct().all()]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Себестоимость"
    ws.append(["Артикул", "Себестоимость"])

    # Заполняем артикулы
    for i, article in enumerate(articles, 2):
        ws[f'A{i}'] = article

    # Сохраняем временный файл
    temp_file = io.BytesIO()
    wb.save(temp_file)
    temp_file.seek(0)

    # Отправляем файл
    file = InputFile(temp_file, filename="шаблон_себестоимость.xlsx")
    await callback.message.answer_document(file, caption="📝 Шаблон для заполнения себестоимости")
    await callback.answer()

async def download_cost_excel_callback(callback: types.CallbackQuery, state: FSMContext):
    session = sessionmaker()(bind=engine)
    try:
        async with state.proxy() as data:
            shop_id = data['shop']['id']
        
        costs = session.query(ProductCost).filter(ProductCost.shop_id == shop_id).all()
        
        if not costs:
            await callback.answer("❌ Нет данных для выгрузки", show_alert=True)
            return
        
        # Создаем Excel файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Себестоимость"
        ws.append(["Артикул", "Себестоимость"])
        
        for cost in costs:
            ws.append([cost.article, cost.cost])
        
        # Сохраняем в BytesIO
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        # Отправляем файл
        file = InputFile(file_stream, filename="себестоимость_артикулов.xlsx")
        await callback.message.answer_document(file, caption="📊 Текущие данные по себестоимости")
    except Exception as e:
        logger.error(f"Ошибка выгрузки себестоимости: {e}")
        await callback.answer("❌ Произошла ошибка при выгрузке данных")
    finally:
        session.close()
        await callback.answer()

async def upload_cost_excel_callback(callback: types.CallbackQuery):
    await callback.message.answer(
        "️ <b>Загрузка данных себестоимости</b>\n\n"
        "Пожалуйста, отправьте Excel-файл в формате:\n"
        "Колонка A: Артикул\n"
        "Колонка B: Себестоимость\n\n"
        "<i>Вы можете скачать шаблон для заполнения</i>"
    )
    await SettingsStates.waiting_for_cost_file.set()

async def process_cost_file(message: types.Message, state: FSMContext):
    if not message.document:
        await message.answer("❌ Пожалуйста, отправьте файл в формате Excel")
        return
    
    if not message.document.file_name.endswith(('.xlsx', '.xls')):
        await message.answer("❌ Неверный формат файла. Отправьте файл Excel (.xlsx или .xls)")
        return
    
    session = sessionmaker()(bind=engine)
    try:
        async with state.proxy() as data:
            shop_id = data['shop']['id']
        
        # Скачиваем файл
        file_id = message.document.file_id
        file = await message.bot.get_file(file_id)
        file_path = file.file_path
        downloaded_file = await message.bot.download_file(file_path)
        
        # Парсим Excel
        wb = openpyxl.load_workbook(io.BytesIO(downloaded_file.read()))
        ws = wb.active
        
        # Собираем данные
        cost_data = {}
        for row in ws.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовок
            if row[0] and row[1]:
                article = row[0]
                try:
                    cost = row[1]
                    cost_data[article] = cost
                except (ValueError, TypeError):
                    continue
        
        if not cost_data:
            await message.answer("❌ В файле не найдены корректные данные")
            return
        
        # Обновляем базу данных
        existing_costs = {cost.article: cost for cost in session.query(ProductCost)
            .filter(ProductCost.shop_id == shop_id)
            .filter(ProductCost.article.in_(cost_data.keys()))
            .all()}
        
        for article, cost_value in cost_data.items():
            if article in existing_costs:
                existing_costs[str(article)].cost = cost_value
            else:
                new_cost = ProductCost(shop_id=shop_id, article=article, cost=cost_value)
                session.add(new_cost)
        
        session.commit()
        await message.answer(f"✅ Успешно обработано: {len(cost_data)} записей")
    except Exception as e:
        logger.error(f"Ошибка обработки файла: {e}")
        await message.answer("❌ Произошла ошибка при обработке файла. Проверьте формат.")
    finally:
        session.close()
        await SettingsStates.product_cost.set()
        await product_cost_callback_helper(message)

# Регулярные затраты
async def regular_expenses_callback(callback: types.CallbackQuery, state: FSMContext):
    session = sessionmaker()(bind=engine)
    try:
        async with state.proxy() as data:
            shop_id = data['shop']['id']
        
        # Подсчет количества расходов
        count = session.query(RegularExpense).filter(
            RegularExpense.shop_id == shop_id
        ).count()
        
        keyboard = InlineKeyboardMarkup(row_width=1)
        keyboard.add(
            InlineKeyboardButton("➕ Добавить расход", callback_data="add_regular_expense"),
            InlineKeyboardButton(f"📋 Список расходов ({count})", callback_data="list_regular_expenses"),
            InlineKeyboardButton("🔙 Назад", callback_data="back_to_settings")
        )
        
        await SettingsStates.regular_expenses.set()
        await callback.message.edit_text(
            " <b>Регулярные затраты</b>\n\nЗдесь Вы можете внести свои регулярные операционные затраты бизнеса, такие как зарплаты сотрудникам, расходные материалы и т.д.",
            reply_markup=keyboard
        )
    finally:
        session.close()

async def add_regular_expense_callback(callback: types.CallbackQuery):
    await callback.message.answer(
        "🔅 <b>Добавление регулярного расхода</b>\n\n"
        "Введите сумму расхода в рублях:"
    )
    await SettingsStates.waiting_for_regular_amount.set()

async def process_regular_amount(message: types.Message, state: FSMContext):
    try:
        amount = float(message.text)
        if amount <= 0:
            raise ValueError
        
        async with state.proxy() as data:
            data['regular_amount'] = amount
        
        await message.answer("📝 Введите описание расхода:")
        await SettingsStates.waiting_for_regular_description.set()
    except ValueError:
        await message.answer("❌ Неверный формат суммы. Введите положительное число:")

async def process_regular_description(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['regular_description'] = message.text
    
    await message.answer(
        "⏱ <b>Выберите периодичность расхода:</b>",
        reply_markup=regular_expense_frequency_keyboard()
    )
    await SettingsStates.waiting_for_regular_frequency.set()

async def set_regular_frequency(callback: types.CallbackQuery, state: FSMContext):
    frequency = RegularExpenseFrequency(callback.data.split('_')[1])
    
    session = sessionmaker()(bind=engine)
    try:
        async with state.proxy() as data:
            shop_id = data['shop']['id']
            amount = data['regular_amount']
            description = data['regular_description']
        
        # Создаем запись
        expense = RegularExpense(
            shop_id=shop_id,
            amount=amount,
            description=description,
            frequency=frequency
        )
        session.add(expense)
        session.commit()
        
        await callback.answer(f"✅ Регулярный расход добавлен: {frequency.value}")
        await back_to_settings(callback, state)
    except Exception as e:
        logger.error(f"Ошибка добавления расхода: {e}")
        await callback.answer("❌ Произошла ошибка при добавлении расхода")
    finally:
        session.close()

# Разовые затраты
async def one_time_expenses_callback(callback: types.CallbackQuery, state: FSMContext):
    session = sessionmaker()(bind=engine)
    try:
        async with state.proxy() as data:
            shop_id = data['shop']['id']
        count = session.query(OneTimeExpense).filter(
            OneTimeExpense.shop_id == shop_id
        ).count()
        
        keyboard = InlineKeyboardMarkup(row_width=1)
        keyboard.add(
            InlineKeyboardButton("➕ Добавить расход", callback_data="add_one_time_expense"),
            InlineKeyboardButton(f"📋 Список расходов ({count})", callback_data="list_one_time_expenses"),
            InlineKeyboardButton("🔙 Назад", callback_data="back_to_settings")
        )
        
        await SettingsStates.one_time_expenses.set()
        await callback.message.edit_text(
            " <b>Ваши первоначальные вложения</b>\n\nЗдесь Вы можете внести свои первоначальные затраты на запуск своего бизнеса. Туда входят открытие ИП, закупка первой партии товара, фотографии и другие вложения, которые Вы делали из своего кармана, а не из оборотных средств.",
            reply_markup=keyboard
        )
    finally:
        session.close()

async def list_regular_expenses_callback(callback: types.CallbackQuery, state: FSMContext):
    session = sessionmaker()(bind=engine)
    try:
        async with state.proxy() as data:
            shop_id = data['shop']['id']
        
        expenses = session.query(RegularExpense).filter(
            RegularExpense.shop_id == shop_id
        ).order_by(RegularExpense.created_at.desc()).all()
        
        if not expenses:
            await callback.answer("📭 Список регулярных расходов пуст", show_alert=True)
            return
        
        text = "📋 <b>Список регулярных расходов</b>\n\n"
        keyboard = InlineKeyboardMarkup(row_width=1)
        
        for expense in expenses:
            freq_map = {
                RegularExpenseFrequency.DAILY: "ежедневно",
                RegularExpenseFrequency.WEEKLY: "еженедельно",
                RegularExpenseFrequency.MONTHLY: "ежемесячно"
            }
            frequency = freq_map.get(expense.frequency, expense.frequency.value)
            
            text += (
                f"💰 <b>{expense.amount:.2f} руб.</b> {frequency}\n"
                f"📝 {expense.description}\n"
                f"🆔 ID: {expense.id}\n\n"
            )
            
            # Кнопка для удаления
            keyboard.add(InlineKeyboardButton(
                f"❌ Удалить расход {expense.id}",
                callback_data=f"delete_regular_{expense.id}"
            ))
        
        text += "ℹ️ Для удаления выберите расход ниже"
        keyboard.add(InlineKeyboardButton("🔙 Назад", callback_data="regular_expenses"))
        
        await callback.message.edit_text(text, reply_markup=keyboard)
    finally:
        session.close()

# Удаление регулярного расхода
async def delete_regular_expense_callback(callback: types.CallbackQuery, state: FSMContext):
    expense_id = int(callback.data.split('_')[2])
    
    session = sessionmaker()(bind=engine)
    try:
        expense = session.query(RegularExpense).get(expense_id)
        if not expense:
            await callback.answer("❌ Расход не найден", show_alert=True)
            return
        
        session.delete(expense)
        session.commit()
        
        await callback.answer(f"✅ Расход {expense_id} удален!")
        await list_regular_expenses_callback(callback, state)
    except Exception as e:
        logger.error(f"Ошибка удаления расхода: {e}")
        await callback.answer("❌ Произошла ошибка при удалении")
    finally:
        session.close()

# Разовые расходы - список
async def list_one_time_expenses_callback(callback: types.CallbackQuery, state: FSMContext):
    session = sessionmaker()(bind=engine)
    try:
        async with state.proxy() as data:
            shop_id = data['shop']['id']
        
        expenses = session.query(OneTimeExpense).filter(
            OneTimeExpense.shop_id == shop_id
        ).order_by(OneTimeExpense.expense_date.desc()).all()
        
        if not expenses:
            await callback.answer(" Список разовых расходов пуст", show_alert=True)
            return
        
        text = "📋 <b>Список разовых расходов</b>\n\n"
        keyboard = InlineKeyboardMarkup(row_width=1)
        
        for expense in expenses:
            # Форматируем дату
            date_str = expense.expense_date.strftime("%d.%m.%Y")
            
            text += (
                f"💰 <b>{expense.amount:.2f} руб.</b> ({date_str})\n"
                f"📝 {expense.description}\n"
                f"🆔 ID: {expense.id}\n\n"
            )
            
            # Кнопка для удаления
            keyboard.add(InlineKeyboardButton(
                f"❌ Удалить расход {expense.id}",
                callback_data=f"delete_onetime_{expense.id}"
            ))
        
        text += "ℹ️ Для удаления выберите расход ниже"
        keyboard.add(InlineKeyboardButton("🔙 Назад", callback_data="one_time_expenses"))
        
        await callback.message.edit_text(text, reply_markup=keyboard)
    finally:
        session.close()

# Удаление разового расхода
async def delete_onetime_expense_callback(callback: types.CallbackQuery):
    expense_id = int(callback.data.split('_')[2])
    
    session = sessionmaker()(bind=engine)
    try:
        expense = session.query(OneTimeExpense).get(expense_id)
        if not expense:
            await callback.answer("❌ Расход не найден", show_alert=True)
            return
        
        session.delete(expense)
        session.commit()
        
        await callback.answer(f"✅ Расход {expense_id} удален!")
        await list_one_time_expenses_callback(callback, state)
    except Exception as e:
        logger.error(f"Ошибка удаления расхода: {e}")
        await callback.answer("❌ Произошла ошибка при удалении")
    finally:
        session.close()

async def add_one_time_expense_callback(callback: types.CallbackQuery):
    await callback.message.answer(
        "💸 <b>Добавление разового расхода</b>\n\n"
        "Введите сумму расхода в рублях:"
    )
    await SettingsStates.waiting_for_onetime_amount.set()

async def process_onetime_amount(message: types.Message, state: FSMContext):
    try:
        amount = float(message.text)
        if amount <= 0:
            raise ValueError
        
        async with state.proxy() as data:
            data['onetime_amount'] = amount
        
        await message.answer("📝 Введите описание расхода:")
        await SettingsStates.waiting_for_onetime_description.set()
    except ValueError:
        await message.answer("❌ Неверный формат суммы. Введите положительное число:")

async def process_onetime_description(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        shop_id = data['shop']['id']
        amount = data['onetime_amount']
        description = message.text
    session = sessionmaker()(bind=engine)
    try:
        expense = OneTimeExpense(
            shop_id=shop_id,
            amount=amount,
            description=description,
            expense_date=datetime.now()  # Для кнопки "Сегодня"
        )
        session.add(expense)
        session.commit()

        await message.answer("✅ Разовый расход добавлен")
        await back_to_settings_message(message, state)
    except Exception as e:
        logger.error(f"Ошибка добавления расхода: {e}")
        await message.answer("❌ Произошла ошибка при добавлении расхода")
        await message.answer("✅ Разовый расход добавлен")
    finally:
        session.close()

async def set_onetime_date(callback: types.CallbackQuery, state: FSMContext):
    session = sessionmaker()(bind=engine)
    try:
        async with state.proxy() as data:
            shop_id = data['shop']['id']
            amount = data['onetime_amount']
            description = data['onetime_description']
        
        # Создаем запись
        expense = OneTimeExpense(
            shop_id=shop_id,
            amount=amount,
            description=description,
            expense_date=datetime.now()  # Для кнопки "Сегодня"
        )
        session.add(expense)
        session.commit()
        
        await callback.answer("✅ Разовый расход добавлен")
        await back_to_settings(callback, state)
    except Exception as e:
        logger.error(f"Ошибка добавления расхода: {e}")
        await callback.answer("❌ Произошла ошибка при добавлении расхода")
    finally:
        session.close()

async def process_onetime_date(message: types.Message, state: FSMContext):
    try:
        day, month, year = map(int, message.text.split('.'))
        expense_date = datetime(year, month, day)
        
        if expense_date > datetime.now():
            await message.answer("❌ Дата не может быть в будущем. Введите корректную дату:")
            return
        
        session = sessionmaker()(bind=engine)
        try:
            async with state.proxy() as data:
                shop_id = data['shop']['id']
                amount = data['onetime_amount']
                description = data['onetime_description']
            
            # Создаем запись
            expense = OneTimeExpense(
                shop_id=shop_id,
                amount=amount,
                description=description,
                expense_date=expense_date
            )
            session.add(expense)
            session.commit()
            
            await message.answer("✅ Разовый расход добавлен")
            await back_to_settings(message, state)
        except Exception as e:
            logger.error(f"Ошибка добавления расхода: {e}")
            await message.answer("❌ Произошла ошибка при добавлении расхода")
        finally:
            session.close()
    except (ValueError, IndexError):
        await message.answer("❌ Неверный формат даты. Введите в формате ДД.ММ.ГГГГ:")

async def test_reports_callback(callback: types.CallbackQuery, state):
    """Тестовая команда для проверки работы планировщика отчетов"""
    try:
        from tg_bot.services.scheduler import ReportScheduler
        
        # Создаем планировщик для тестирования
        scheduler = ReportScheduler(callback.bot)
        
        # Тестируем еженедельный отчет
        await callback.answer("Тестирую еженедельный отчет...", show_alert=True)
        await scheduler.send_weekly_report()
        
        await callback.message.answer("✅ Еженедельный отчет отправлен!")
        
    except Exception as e:
        logger.error(f"Ошибка при тестировании отчетов: {e}")
        await callback.answer(f"Ошибка при тестировании: {e}", show_alert=True)

# Регистрация обработчиков
def register_settings_handlers(dp):
    # Ежедневные отчёты
    dp.register_callback_query_handler(daily_reports_callback, text="daily_reports", state="*")
    # Тест планировщика отчетов
    dp.register_callback_query_handler(test_reports_callback, text="test_reports", state="*")
    # Налоговая система
    
    # Обработчик для кнопки "Налоговая система" - сразу переход к вводу процента
    dp.register_callback_query_handler(
        set_custom_tax_callback, 
        lambda c: c.data == "tax_custom", 
        state="*"
    )
    
    # Обработчик ввода процента с клавиатуры
    dp.register_message_handler(
        process_tax_percent, 
        state=SettingsStates.waiting_for_tax_percent
    )

    dp.register_callback_query_handler(tax_settings_callback, text="tax_settings", state="*")
    dp.register_callback_query_handler(set_tax_system_callback, lambda c: c.data.startswith("tax_") and c.data != "tax_settings", state=SettingsStates.tax_system)
    
    # Себестоимость артикулов
    dp.register_callback_query_handler(product_cost_callback, text="product_cost", state="*")
    dp.register_callback_query_handler(download_cost_template_callback, text="download_cost_template", state=SettingsStates.product_cost)
    dp.register_callback_query_handler(download_cost_excel_callback, text="download_cost_excel", state=SettingsStates.product_cost)
    dp.register_callback_query_handler(upload_cost_excel_callback, text="upload_cost_excel", state=SettingsStates.product_cost)
    dp.register_message_handler(process_cost_file, content_types=types.ContentType.DOCUMENT, state=SettingsStates.waiting_for_cost_file)
    
    dp.register_callback_query_handler(regular_expenses_callback, text="regular_expenses", state="*")
    dp.register_callback_query_handler(add_regular_expense_callback, text="add_regular_expense", state=SettingsStates.regular_expenses)
    dp.register_message_handler(process_regular_amount, state=SettingsStates.waiting_for_regular_amount)
    dp.register_message_handler(process_regular_description, state=SettingsStates.waiting_for_regular_description)
    dp.register_callback_query_handler(set_regular_frequency, lambda c: c.data.startswith("frequency_"), state=SettingsStates.waiting_for_regular_frequency)
    
    dp.register_callback_query_handler(one_time_expenses_callback, text="one_time_expenses", state="*")
    dp.register_callback_query_handler(add_one_time_expense_callback, text="add_one_time_expense", state=SettingsStates.one_time_expenses)
    dp.register_message_handler(process_onetime_amount, state=SettingsStates.waiting_for_onetime_amount)
    dp.register_message_handler(process_onetime_description, state=SettingsStates.waiting_for_onetime_description)
    dp.register_callback_query_handler(set_onetime_date, text="expense_date_today", state=SettingsStates.waiting_for_onetime_date)
    dp.register_message_handler(process_onetime_date, state=SettingsStates.waiting_for_onetime_date)

    # After start
    dp.register_callback_query_handler(settings_callback, text="settings", state="*")
    dp.register_callback_query_handler(back_to_settings, text="back_to_settings", state="*")

    # Регулярные расходы
    dp.register_callback_query_handler(
        list_regular_expenses_callback, 
        text="list_regular_expenses", 
        state=SettingsStates.regular_expenses
    )
    dp.register_callback_query_handler(
        delete_regular_expense_callback, 
        lambda c: c.data.startswith("delete_regular_"), 
        state="*"
    )
    
    # Разовые расходы
    dp.register_callback_query_handler(
        list_one_time_expenses_callback, 
        text="list_one_time_expenses", 
        state=SettingsStates.one_time_expenses
    )
    dp.register_callback_query_handler(
        delete_onetime_expense_callback, 
        lambda c: c.data.startswith("delete_onetime_"), 
        state="*"
    )
