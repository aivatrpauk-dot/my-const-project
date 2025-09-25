from aiogram import types
import asyncio
from tg_bot.services.wb_api import fetch_report_detail_by_period, fetch_report_detail_by_period_cached
from aiogram.dispatcher import FSMContext, Dispatcher
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, InputFile
from aiogram.utils.exceptions import MessageNotModified
from tg_bot.models import (
    Order,
    Shop,
    sessionmaker,
    engine,
    ProductCost,
    TaxSystemSetting,
    RegularExpense,
    TaxSystemType,
    OneTimeExpense,
    Advertisement,
    Penalty,
User,
    RegularExpenseFrequency
)
from tg_bot.states.analytics_states import AnalyticsStates
from tg_bot.keyboards.analytics_menu import (
    analytics_menu_keyboard,
    period_keyboard,
    period_keyboard2,
)
from tg_bot.services.wb_api import fetch_full_report
from dateutil.relativedelta import relativedelta
from datetime import datetime, timedelta, date
import math
import io
import logging
import openpyxl
import json
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from tg_bot.handlers.pnl import generate_pnl_excel_report  # Импортируем функцию генерации Excel
import os
import pandas as pd
import numpy as np
from collections import defaultdict
from datetime import datetime as dt


logger = logging.getLogger(__name__)


# Главное меню аналитики
async def analytics_callback(callback: types.CallbackQuery):
    text = (
        "📈 <b>Аналитика и рекомендации</b>\n\n"
        "Здесь вы можете получить детальную аналитику по вашему бизнесу на Wildberries. "
        "Выберите интересующий раздел:"
    )
    keyboard = analytics_menu_keyboard()
    await callback.message.edit_text(text, reply_markup=keyboard)


# Обработчики для подменю
PROFITABILITY_LEVELS = [
    {
        "min": -float("inf"),
        "max": 20,
        "name": "⚠️ Низкая доходность",
        "characteristics": "Плохая рентабельность, высокие риски или низкая маржинальность.",
        "reasons": "Высокая конкуренция, большие расходы на логистику/хранение, низкие наценки.",
        "conclusion": "Такой бизнес невыгоден, нужно пересматривать модель.",
        "recommendations": [
            "Срочно пересмотрите ценовую политику и себестоимость.",
            "Ищите более выгодных поставщиков или сокращайте логистические издержки.",
            "Проверьте скрытые расходы (хранение, возвраты, реклама) и оптимизируйте их.",
            "Если рост невозможен – рассмотрите закрытие или смену ниши.",
        ],
        "action": "Оптимизировать или уходить",
    },
    {
        "min": 20,
        "max": 40,
        "name": "⚠️ Ниже среднего",
        "characteristics": "Минимально приемлемая рентабельность, но требует оптимизации.",
        "reasons": "Средняя конкуренция, умеренные издержки.",
        "conclusion": "Высокий риск уйти в ноль или минус из-за внешних факторов.",
        "recommendations": [
            "Увеличивайте маржу через улучшение упаковки, допродажи или брендинг.",
            "Автоматизируйте процессы для снижения операционных затрат.",
            "Тестируйте новые рекламные каналы для увеличения продаж.",
            "Анализируйте конкурентов на предмет более выгодных товаров.",
        ],
        "action": "Улучшать и тестировать другие товары",
    },
    {
        "min": 40,
        "max": 60,
        "name": "✅ Средняя доходность",
        "characteristics": "Нормальный уровень для стабильного бизнеса.",
        "reasons": "Хороший спрос, грамотное ценообразование, контроль затрат.",
        "conclusion": "Устойчивый бизнес, можно масштабировать.",
        "recommendations": [
            "Фокусируйтесь на стабильности: контролируйте качество и сервис.",
            "Расширяйте ассортимент в нише для увеличения среднего чека.",
            "Инвестируйте в лояльность клиентов (отзывы, рассылки).",
            "Тестируйте смежные ниши с более высокой маржой.",
        ],
        "action": "Закрепляться и расти",
    },
    {
        "min": 60,
        "max": 100,
        "name": "🔥 Высокая доходность",
        "characteristics": "Очень хорошая рентабельность, перспективный бизнес.",
        "reasons": "Уникальный товар, низкая конкуренция, эффективные рекламные каналы.",
        "conclusion": "Отличный результат, стоит вкладывать больше ресурсов.",
        "recommendations": [
            "Активно масштабируйте: выходите на новые маркетплейсы или рынки.",
            "Усиливайте бренд и работайте с повторными продажами.",
            "Диверсифицируйте поставщиков для снижения рисков.",
            "Инвестируйте часть прибыли в новые высокомаржинальные товары.",
        ],
        "action": "Масштабировать и защищать",
    },
    {
        "min": 100,
        "max": float("inf"),
        "name": "✨ Премиальная доходность",
        "characteristics": "Высокомаржинальный бизнес, часто нишевый.",
        "reasons": "Эксклюзивные товары, VIP-сегмент, отсутствие прямых аналогов.",
        "conclusion": "Редкий и ценный кейс, требует защиты позиций.",
        "recommendations": [
            "Укрепляйте эксклюзивность через товарный знак и уникальные условия с поставщиками.",
            "Создавайте финансовую подушку безопасности.",
            "Масштабируйте до точки максимальной эффективности.",
            "Мониторьте динамику прибыли и будьте готовы к поиску новых товаров.",
        ],
        "action": "Укреплять позиции или выжимать все соки",
    },
]


def get_profitability_level(profitability):
    """Определение уровня доходности по проценту рентабельности"""
    for level in PROFITABILITY_LEVELS:
        if level["min"] <= profitability < level["max"]:
            return level
    return PROFITABILITY_LEVELS[0]  # По умолчанию низкая доходность


async def profitability_estimation_callback(
    callback: types.CallbackQuery, state: FSMContext
):
    """Обработчик оценки доходности"""
    # Проверяем выбран ли магазин
    async with state.proxy() as data:
        if "shop" not in data:
            await callback.answer("❌ Сначала выберите магазин", show_alert=True)
            return

    # Сохраняем контекст для пагинации
    async with state.proxy() as data:
        data["analytics_type"] = "profitability"
        data["article_page"] = 0

    await show_articles_page(callback, state)


async def calculate_profitability_for_article(article, shop_id, api_token):
    """Расчет доходности для конкретного артикула"""
    end_date = datetime.utcnow()
    start_date = end_date - timedelta(days=30)

    # Получаем отчет за последний месяц
    report = await fetch_full_report(api_token, start_date, end_date)
    if not report:
        return None

    print("FLAG0")
    # for i in report:
    # print(i['nm_id'])
    # Фильтруем данные по выбранному артикулу
    article_data = [item for item in report if item.get("nm_id") == article]
    print("FLAG0.5")
    print(article_data)
    if not article_data:
        return None
    print("FLAG1")
    # Рассчитываем показатели
    total_quantity = sum(item.get("quantity", 0) for item in article_data)
    total_revenue = sum(
        item.get("retail_price_withdisc_rub", 0) for item in article_data
    )
    total_commission = sum(
        item.get("ppvz_sales_commission", 0)
        + item.get("ppvz_vw", 0)
        + item.get("ppvz_vw_nds", 0)
        for item in article_data
    )
    print("FLAG2")

    # Рассчитываем общие расходы для распределения
    total_logistics = sum(item.get("delivery_rub", 0) for item in report)
    total_storage = sum(item.get("storage_fee", 0) for item in report)
    total_revenue_all = sum(item.get("retail_price_withdisc_rub", 0) for item in report)

    # Распределяем логистику и хранение пропорционально выручке
    logistics_share = (
        total_logistics * (total_revenue / total_revenue_all)
        if total_revenue_all
        else 0
    )
    storage_share = (
        total_storage * (total_revenue / total_revenue_all) if total_revenue_all else 0
    )
    print("FLAG3")
    # Получаем себестоимость
    session = sessionmaker()(bind=engine)
    try:
        product_cost = (
            session.query(ProductCost)
            .filter(ProductCost.shop_id == shop_id, ProductCost.article == article)
            .first()
        )
        cost_per_item = product_cost.cost if product_cost else 0
    finally:
        session.close()

    total_cost = cost_per_item * total_quantity
    total_expenses = total_commission + logistics_share + storage_share + total_cost

    # Рассчитываем прибыль и рентабельность
    net_profit = total_revenue - total_expenses
    profitability = (net_profit / total_revenue) * 100 if total_revenue else 0

    return {
        "revenue": total_revenue,
        "cost": total_cost,
        "commission": total_commission,
        "logistics": logistics_share,
        "storage": storage_share,
        "expenses": total_expenses,
        "net_profit": net_profit,
        "profitability": profitability,
        "quantity": total_quantity,
        "cost_per_item": cost_per_item,
    }
def get_comm(comission, category):
    for cat in comission["report"]:
        if cat["parentName"] == category:
            return cat["paidStorageKgvp"]

async def show_profitability_report(
    callback: types.CallbackQuery, article, state: FSMContext
):
    """Показать отчет по доходности для артикула"""
    async with state.proxy() as data:
        shop_id = data["shop"]["id"]
        shop_name = data["shop"]["name"] or f"Магазин {shop_id}"
        api_token = data["shop"]["api_token"]

    # Показываем сообщение о загрузке
    await callback.message.edit_text(
        f"📊 <b>Расчет доходности для артикула {article}</b>\n\n"
        f"Магазин: {shop_name}\n"
        "Период: последний месяц\n\n"
        "Подождите, идет расчет..."
    )

    # Рассчитываем показатели
    metrics = await calculate_profitability_for_article(article, shop_id, api_token)

    if not metrics:
        await callback.message.edit_text(
            f"❌ <b>Не удалось рассчитать доходность для артикула {article}</b>\n\n"
            "Возможные причины:\n"
            "1. Нет данных о продажах за последний месяц\n"
            "2. Не загружена себестоимость товара\n"
            "3. Проблемы с подключением к WB API"
        )
        return

    # Определяем уровень доходности
    profitability = metrics["profitability"]
    level = get_profitability_level(profitability)

    # Форматируем отчет
    text = (
        f"📊 <b>Оценка доходности: {level['name']}</b>\n\n"
        f"<b>Артикул:</b> {article}\n"
        f"<b>Магазин:</b> {shop_name}\n"
        f"<b>Период:</b> последний месяц\n\n"
        "<u>Финансовые показатели:</u>\n"
        f"💰 Выручка: {metrics['revenue']:.2f} руб.\n"
        f"📦 Продано: {metrics['quantity']} шт.\n"
        f"🏷️ Себестоимость: {metrics['cost_per_item']:.2f} руб./шт. (Итого: {metrics['cost']:.2f} руб.)\n"
        f"📊 Комиссии WB: {metrics['commission']:.2f} руб.\n"
        f"🚚 Логистика: {metrics['logistics']:.2f} руб.\n"
        f"🏭 Хранение: {metrics['storage']:.2f} руб.\n"
        f"💵 Чистая прибыль: {metrics['net_profit']:.2f} руб.\n"
        f"📈 Рентабельность: <b>{profitability:.1f}%</b>\n\n"
        f"<u>Характеристика:</u>\n{level['characteristics']}\n\n"
        f"<u>Основные причины:</u>\n{level['reasons']}\n\n"
        f"<u>Вывод:</u>\n{level['conclusion']}\n\n"
        "<u>Рекомендации:</u>\n"
    )

    # Добавляем рекомендации
    for i, recommendation in enumerate(level["recommendations"]):
        text += f"{i+1}. {recommendation}\n"

    text += f"\n<u>Действие:</u>\n🚀 <b>{level['action']}</b>"

    keyboard = InlineKeyboardMarkup()
    keyboard.add(
        InlineKeyboardButton(
            "🔙 К выбору артикула", callback_data="profitability_estimation"
        )
    )
    keyboard.add(InlineKeyboardButton("📊 В меню аналитики", callback_data="analytics"))

    await callback.message.edit_text(text, reply_markup=keyboard)


async def get_top_profitable_products(api_token: str, shop_id: int):
    end_date = datetime.utcnow()
    start_date = end_date - timedelta(days=30)
    report = await fetch_full_report(api_token, start_date, end_date)
    if not report:
        return []

    session = sessionmaker()(bind=engine)
    try:
        products = {}

        for item in report:
            article = item.get("sa_name")
            if not article:
                continue

            quantity = item.get("quantity", 0)
            revenue = item.get("retail_price_withdisc_rub", 0) * quantity
            commission = (
                item.get("ppvz_sales_commission", 0)
                + item.get("ppvz_vw", 0)
                + item.get("ppvz_vw_nds", 0)
            )
            logistics = item.get("delivery_rub", 0) / len(report) * quantity + item.get("rebill_logistic_cost", 0) / len(report) * quantity
            storage = item.get("storage_fee", 0) / len(report) * quantity

            product_cost = (
                session.query(ProductCost)
                .filter(ProductCost.shop_id == shop_id, ProductCost.article == article)
                .first()
            )
            cost = product_cost.cost if product_cost else 0

            profit = revenue - (cost * quantity + commission + logistics + storage)

            if article not in products:
                products[article] = {
                    "revenue": 0,
                    "cost": cost,
                    "quantity": 0,
                    "profit": 0,
                }

            products[article]["revenue"] += revenue
            products[article]["quantity"] += quantity
            products[article]["profit"] += profit

        sorted_products = sorted(
            products.items(), key=lambda x: x[1]["profit"], reverse=True
        )

        return sorted_products[:5]

    except Exception as e:
        logger.error(f"Ошибка расчета топ-5 товаров: {e}")
        return []
    finally:
        session.close()


async def top5_products_callback(callback: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        if "shop" not in data:
            await callback.answer("❌ Сначала выберите магазин", show_alert=True)
            return

        shop_id = data["shop"]["id"]
        shop_name = data["shop"]["name"] or f"Магазин {shop_id}"
        api_token = data["shop"]["api_token"]

    await callback.message.edit_text(
        "⏳ <b>Расчет топ-5 самых прибыльных товаров</b>\n\n"
        f"Магазин: {shop_name}\n"
        "Период: последний месяц\n\n"
        "Подождите, идет расчет..."
    )

    top_products = await get_top_profitable_products(api_token, shop_id)

    if not top_products:
        await callback.message.edit_text(
            "❌ <b>Не удалось получить данные</b>\n\n"
            "Проверьте, что:\n"
            "1. У вас есть подключение к интернету\n"
            "2. API-токен WB действителен\n"
            "3. Загружены данные себестоимости"
        )
        return

    text = (
        f"🏆 <b>Топ-5 самых прибыльных товаров</b>\n\n"
        f"Магазин: {shop_name}\n"
        "Период: последний месяц\n\n"
    )

    for i, (article, data) in enumerate(top_products):
        profit = data["profit"]
        revenue = data["revenue"]
        quantity = data["quantity"]
        cost = data["cost"]

        text += (
            f"{i+1}. <b>{article}</b>\n"
            f"   Прибыль: {profit:.2f} руб.\n"
            f"   Выручка: {revenue:.2f} руб.\n"
            f"   Продано: {quantity} шт.\n"
            f"   Себестоимость: {cost:.2f} руб./шт.\n\n"
        )

    text += "<i>Примечание: расчет включает себестоимость, комиссии, логистику и хранение</i>"

    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("🔙 Назад", callback_data="analytics"))

    await callback.message.edit_text(text, reply_markup=keyboard)


async def what_if_simulator_callback(callback: types.CallbackQuery, state: FSMContext):
    # Проверяем выбран ли магазин
    async with state.proxy() as data:
        if "shop" not in data:
            await callback.answer("❌ Сначала выберите магазин", show_alert=True)
            return
        shop_id = data["shop"]["id"]

    # Сохраняем контекст для пагинации
    async with state.proxy() as data:
        data["analytics_type"] = "what_if"
        data["article_page"] = 0

    await show_articles_page(callback, state)


async def show_articles_page(callback: types.CallbackQuery, state: FSMContext):
    session = sessionmaker()(bind=engine)
    try:
        async with state.proxy() as data:
            shop_id = data["shop"]["id"]
            page = data["article_page"]
            analytics_type = data["analytics_type"]

        articles = (
            session.query(Order.nmId).filter(Order.shop_id == shop_id).distinct().all()
        )

        articles = [art[0] for art in articles]

        if not articles:
            await callback.answer("❌ Нет данных по артикулам", show_alert=True)
            return
        items_per_page = 7
        total_pages = math.ceil(len(articles) / items_per_page)
        start_idx = page * items_per_page
        page_articles = articles[start_idx : start_idx + items_per_page]

        title = (
            "📊 Оценка доходности"
            if analytics_type == "profitability"
            else "🔮 Симулятор «А что если?»"
        )
        text = f"{title}\n\nВыберите артикул (страница {page + 1}/{total_pages}):"

        keyboard = InlineKeyboardMarkup(row_width=1)

        for article in page_articles:
            keyboard.add(
                InlineKeyboardButton(article, callback_data=f"select_article_{article}")
            )

        pagination_row = []
        if page > 0:
            pagination_row.append(
                InlineKeyboardButton("⬅️ Назад", callback_data="prev_articles_page")
            )
        if start_idx + items_per_page < len(articles):
            pagination_row.append(
                InlineKeyboardButton("Вперед ➡️", callback_data="next_articles_page")
            )

        if pagination_row:
            keyboard.row(*pagination_row)

        keyboard.add(InlineKeyboardButton("🔙 Назад", callback_data="back_to_analytics"))
        await callback.message.delete()
        await callback.message.answer(text, reply_markup=keyboard)
        await AnalyticsStates.waiting_for_article.set()
    finally:
        session.close()


async def handle_articles_pagination(callback: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        page = data["article_page"]
        if callback.data == "prev_articles_page":
            data["article_page"] = max(0, page - 1)
        else:
            data["article_page"] = page + 1

    await show_articles_page(callback, state)


async def select_article_callback(callback: types.CallbackQuery, state: FSMContext):
    article = callback.data.split("_", 2)[2]

    async with state.proxy() as data:
        analytics_type = data["analytics_type"]
        shop_id = data["shop"]["id"]
        data["selected_article"] = article

    if analytics_type == "profitability":
        await show_profitability_report(callback, int(article), state)
    else:
        await callback.message.edit_text(
            "🔮 <b>Симулятор «А что если?»</b>\n\n"
            f"Выбран артикул: <b>{article}</b>\n\n"
            "Введите новую цену и новую себестоимость через запятую.\n"
            "Формат: <code>цена, себестоимость</code>\n"
            "Например: <code>1200, 800</code>"
        )
        await AnalyticsStates.waiting_for_price_and_cost.set()


async def what_if_simulator_callback(callback: types.CallbackQuery, state: FSMContext):
    """Обработчик симулятора сценариев"""
    # Проверяем выбран ли магазин
    async with state.proxy() as data:
        if "shop" not in data:
            await callback.answer("❌ Сначала выберите магазин", show_alert=True)
            return

    # Сохраняем контекст для пагинации
    async with state.proxy() as data:
        data["analytics_type"] = "what_if"
        data["article_page"] = 0

    await show_articles_page(callback, state)


async def process_price_and_cost(message: types.Message, state: FSMContext):
    """Обработка ввода цены и себестоимости"""
    try:
        # Пытаемся разобрать ввод
        input_text = message.text.strip()

        # Проверяем два формата: через запятую и через пробел
        if "," in input_text:
            parts = input_text.split(",")
        else:
            parts = input_text.split()

        if len(parts) != 2:
            raise ValueError

        new_price = float(parts[0].strip())
        new_cost = float(parts[1].strip())

        async with state.proxy() as data:
            article = data["selected_article"]
            shop_id = data["shop"]["id"]
            shop_name = data["shop"]["name"] or f"Магазин {shop_id}"
            api_token = data["shop"]["api_token"]

        # Получаем исторические данные
        end_date = datetime.utcnow()
        start_date = end_date - timedelta(days=30)
        report = await fetch_full_report(api_token, start_date, end_date)

        if not report:
            await message.answer(
                "❌ Не удалось получить исторические данные для расчета"
            )
            return

        # Фильтруем данные по выбранному артикулу
        article_data = [item for item in report if item.get("sa_name") == article]

        if not article_data:
            await message.answer(
                f"❌ Нет данных по артикулу {article} за последний месяц"
            )
            return

        # Рассчитываем текущие показатели
        current_quantity = sum(item.get("quantity", 0) for item in article_data)
        current_revenue = sum(
            item.get("retail_price_withdisc_rub", 0) for item in article_data
        )
        current_commission = sum(
            item.get("ppvz_sales_commission", 0)
            + item.get("ppvz_vw", 0)
            + item.get("ppvz_vw_nds", 0)
            for item in article_data
        )

        # Получаем текущую себестоимость
        session = sessionmaker()(bind=engine)
        try:
            product_cost = (
                session.query(ProductCost)
                .filter(
                    ProductCost.shop_id == shop_id, ProductCost.article == str(article)
                )
                .first()
            )
            current_cost = product_cost.cost if product_cost else 0
        finally:
            session.close()

        current_profit = (
            current_revenue - current_commission - (current_cost * current_quantity)
        )

        # Рассчитываем прогноз
        forecast_revenue = new_price * current_quantity
        forecast_profit = (
            forecast_revenue - current_commission - (new_cost * current_quantity)
        )

        # Формируем результат
        text = (
            f"🔮 <b>Симулятор «А что если?» для артикула {article}</b>\n\n"
            f"<b>Исторические данные (за последний месяц):</b>\n"
            f"📦 Продано: {current_quantity} шт.\n"
            f"💰 Выручка: {current_revenue:.2f} руб.\n"
            f"💵 Прибыль: {current_profit:.2f} руб.\n"
            f"🏷️ Текущая цена: {current_revenue / current_quantity:.2f} руб./шт.\n"
            f"📊 Текущая себестоимость: {current_cost:.2f} руб./шт.\n\n"
            f"<b>Прогноз при новых параметрах:</b>\n"
            f"🆕 Новая цена: {new_price:.2f} руб./шт.\n"
            f"🆕 Новая себестоимость: {new_cost:.2f} руб./шт.\n"
            f"📈 Прогнозируемая выручка: {forecast_revenue:.2f} руб.\n"
            f"📊 Прогнозируемая прибыль: {forecast_profit:.2f} руб.\n\n"
            f"<b>Изменение:</b>\n"
            f"💰 Выручка: {forecast_revenue - current_revenue:+.2f} руб. "
            f"({(forecast_revenue / current_revenue - 1) * 100 if current_revenue else 0:+.1f}%)\n"
            f"💵 Прибыль: {forecast_profit - current_profit:+.2f} руб. "
            f"({(forecast_profit / current_profit - 1) * 100 if current_profit else 0:+.1f}%)\n\n"
            "<i>Примечание: прогноз основан на историческом количестве продаж без учета изменения спроса</i>"
        )

        keyboard = InlineKeyboardMarkup()
        keyboard.add(
            InlineKeyboardButton("🔄 Новый расчет", callback_data="what_if_simulator")
        )
        keyboard.add(
            InlineKeyboardButton("🔙 В меню аналитики", callback_data="analytics")
        )

        await message.answer(text, reply_markup=keyboard)
        await state.finish()

    except (ValueError, IndexError):
        await message.answer(
            "❌ Неверный формат. Пожалуйста, введите цену и себестоимость через запятую или пробел.\n"
            "Пример: <code>1200, 800</code> или <code>1200 800</code>"
        )
    except Exception as e:
        logger.error(f"Ошибка в симуляторе: {e}")
        await message.answer(
            "❌ Произошла ошибка при расчете прогноза. Попробуйте позже."
        )
        await state.finish()


async def product_analytics_callback(callback: types.CallbackQuery, state: FSMContext,start_date,end_date):
    print("product_analytics_callback")
    
    # Получаем данные состояния
    data = await state.get_data()
    shop_data = data.get("shop")
    
    if not shop_data:
        print("shop not in data")
        await callback.answer("❌ Сначала выберите магазин", show_alert=True)
        return

    shop_id = shop_data["id"]
    shop_name = shop_data.get("name") or f"Магазин {shop_id}"
    api_token = shop_data["api_token"]

    # Показываем сообщение о загрузкеc
    await callback.message.delete()
    message2 = await callback.message.answer(
        "<b>Генерация отчёта по товарам</b>\n\n"
        f"Магазин: {shop_name}\n"
        "Подождите, идет сбор и обработка данных..."
    )

    # Получаем данные и генерируем отчет
    try:
        # Получаем данные из API
        loop = asyncio.get_event_loop()
        full_data = await loop.run_in_executor(
            None,
            fetch_report_detail_by_period,
            api_token,
            start_date,
            end_date
        )
        wb = await generate_product_analytics_report(api_token, shop_id, start_date, end_date, full_data)
        if not wb:
            await message2.edit_text(
                "❌ <b>Не удалось сгенерировать отчет</b>\n\n"
                "Возможные причины:\n"
                "1. Нет данных о продажах\n"
                "2. Проблемы с подключением к WB API\n"
                "3. Отсутствуют данные себестоимости"
            )
            return

        # Сохраняем в буфер
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        # Отправляем файл
        await message2.delete()
        file = InputFile(file_stream, filename=f"отчет по товарам .xlsx")
        await callback.message.answer_document(
            file,
            caption=f"Товарная аналитика за последний месяц\nМагазин: {shop_name}",
        )

    except Exception as e:
        logger.error(f"Ошибка генерации отчета: {e}")
        await callback.message.edit_text(
            "❌ Произошла ошибка при генерации отчета. Попробуйте позже."
        )

#Вызывает календарь когда тригерится функция с эксель отчётом
async def start_analytics_report(callback: types.CallbackQuery, state: FSMContext):
    # Сохраняем, что после выбора периода нужно вызвать generate_product_analytics_report
    await state.update_data(report_type="product_analytics")
    await custom_period_callback(callback, state)  # Показываем выбор периода

async def generate_product_analytics_report(api_token: str, shop_id: int, start_date, end_date, full_data):
    """Генерация Excel-отчета с товарной аналитикой"""
    session = sessionmaker()(bind=engine)
    try:
        # Проверяем структуру данных
        if isinstance(full_data, dict):
            report_data = full_data.get('finance', [])
            orders = full_data.get('orders', [])
            sales = full_data.get('sales', [])
        else:
            # Если full_data не словарь, считаем что это finance данные
            report_data = full_data if isinstance(full_data, list) else []
            orders = []
            sales = []
        
        # Используем выбранный период напрямую
        df_orders = pd.DataFrame(orders)
        df_sales = pd.DataFrame(sales)
        df_fin = pd.DataFrame(report_data)

        # day‑столбец
        if not df_orders.empty:
            df_orders["day"] = pd.to_datetime(df_orders["date"]).dt.date
        else:
            df_orders["day"] = pd.to_datetime("2025-01-01").date()
            
        if not df_sales.empty:
            df_sales["day"] = pd.to_datetime(df_sales["date"]).dt.date
        else:
            df_sales["day"] = pd.to_datetime("2025-01-01").date()
        
        # Проверяем наличие поля rr_dt в финансовых данных
        if not df_fin.empty:
            if "rr_dt" in df_fin.columns:
                df_fin["day"] = pd.to_datetime(df_fin["rr_dt"]).dt.date
            elif "date" in df_fin.columns:
                df_fin["day"] = pd.to_datetime(df_fin["date"]).dt.date
            else:
                # Если нет поля даты, создаем пустой столбец day
                df_fin["day"] = pd.to_datetime("2025-01-01").date()
        else:
            df_fin["day"] = pd.to_datetime("2025-01-01").date()

        # 4.1 Orders
        if not df_orders.empty:
            grp_orders = (
                df_orders.groupby("day")
                        .agg(
                            order_sum = ("priceWithDisc", "sum"),
                            ordered_qty = ("srid", "size"),
                            orders_cnt = ("srid", "nunique")
                        )
            )
        else:
            # Создаем пустой DataFrame с нужными столбцами
            grp_orders = pd.DataFrame(columns=['order_sum', 'ordered_qty', 'orders_cnt'])

        # 4.2 Sales (только продажи, не возвраты)
        if not df_sales.empty:
            sales_mask = df_sales["saleID"].str.startswith("S", na=False)
            grp_sales = (
                df_sales[sales_mask]
                .groupby("day") 
                .agg(
                    sales_sum = ("priceWithDisc", "sum"),
                    sold_qty = ("saleID", "size"),
                    for_pay = ("forPay", "sum")
                )
            )
        else:
            # Создаем пустой DataFrame с нужными столбцами
            grp_sales = pd.DataFrame(columns=['sales_sum', 'sold_qty', 'for_pay'])

        # 4.3 Finance — logistica = только delivery_rub
        if not df_fin.empty:
            grp_fin = (
                df_fin.groupby("day")
                    .agg(
                        delivery_cost = ("delivery_rub", "sum"),
                        storage = ("storage_fee", "sum"),
                        penalty = ("penalty", "sum"),
                        acceptance = ("acceptance", "sum"),
                        pay_for_goods = ("ppvz_for_pay", "sum")
                    )
                    .apply(pd.to_numeric, errors="coerce")
            )
        else:
            # Создаем пустой DataFrame с нужными столбцами
            grp_fin = pd.DataFrame(columns=['delivery_cost', 'storage', 'penalty', 'acceptance', 'pay_for_goods'])

        # ────────────────────────────────────────────────────────────────────
        # 5. Объединяем и рассчитываем
        # ────────────────────────────────────────────────────────────────────
        all_daily = (
            grp_orders
            .join(grp_sales, how="outer")
            .join(grp_fin, how="outer")
            .fillna(0)
            .infer_objects(copy=False)
        )

        all_daily["buyout_pct"] = (
            all_daily["sold_qty"] / all_daily["ordered_qty"]
        ).fillna(0)

        all_daily["total_to_pay"] = (
            all_daily["pay_for_goods"]
            - all_daily[["delivery_cost", "storage", "penalty", "acceptance"]].sum(axis=1)
        )

        # Используем все данные за выбранный период
        daily = all_daily
        
        # Отладочная информация
        logger.info(f"Период: {start_date.date()} - {end_date.date()}")
        logger.info(f"Всего записей в daily: {len(daily)}")
        if not daily.empty:
            logger.info(f"Индексы daily: {daily.index.tolist()}")

        if not report_data:
            print("no data from API")
            return None

        # Создаем Excel-книгу
        try:
            wb = load_workbook("template.xlsx")
        except FileNotFoundError:
            wb = openpyxl.Workbook()
        
        ws = wb.active
        ws.title = "Товарная аналитика"



        # Получаем рекламу и штрафы из базы данных
        advert = sum(
            i.amount for i in session.query(Advertisement)
            .filter(Advertisement.shop_id == shop_id)
            .filter(Advertisement.date >= start_date)
            .filter(Advertisement.date <= end_date)
        )

        stops = sum(
            i.amount for i in session.query(Penalty)
            .filter(Penalty.shop_id == shop_id)
            .filter(Penalty.date >= start_date)
            .filter(Penalty.date <= end_date)
        )

        # Получаем настройки налоговой системы
        tax_setting = session.query(TaxSystemSetting).filter(TaxSystemSetting.shop_id == shop_id).first()
        
        # Рассчитываем налог
        revenue = daily["sales_sum"].sum()
        if tax_setting:
            if tax_setting.tax_system == TaxSystemType.USN_6:
                tax_rate = 0.06
                print("tax_rate = ", tax_rate)
            elif tax_setting.tax_system == TaxSystemType.NO_TAX:
                tax_rate = 0.0
                print("tax_rate = ", tax_rate)
            elif tax_setting.tax_system == TaxSystemType.CUSTOM:
                tax_rate = tax_setting.custom_percent / 100
                print("tax_rate = ", tax_rate)
            else:
                tax_rate = 0.0
                print("пошло в else")
        else:
            tax_rate = 0.0
            
        tax = revenue * tax_rate
        
        # Себестоимость (как в analytics.py)
        cost_of_goods = 0
        articles = {}
        
        # Собираем артикулы для расчета себестоимости
        for item in report_data:
            if not isinstance(item, dict):
                continue
            article = item.get("nm_id")
            quantity = item.get("quantity", 0)
            if article and quantity:
                if article not in articles:
                    articles[article] = 0
                articles[article] += quantity

        # Рассчитываем себестоимость
        for article, quantity in articles.items():
            try:
                supp_article = session.query(Order).filter(Order.nmId == int(article)).first().supplierArticle
                product_cost = (
                    session.query(ProductCost)
                    .filter(ProductCost.shop_id == shop_id, ProductCost.article == supp_article)
                    .first()
                )
                if product_cost:
                    cost_of_goods += product_cost.cost * quantity
            except:
                pass

        # Рассчитываем регулярные расходы за период
        regular_expenses = 0
        days_in_period = (end_date - start_date).days + 1
        for expense in session.query(RegularExpense).filter(
                RegularExpense.shop_id == shop_id
        ):
            if expense.frequency == RegularExpenseFrequency.DAILY:
                regular_expenses += expense.amount * days_in_period
            elif expense.frequency == RegularExpenseFrequency.WEEKLY:
                regular_expenses += expense.amount * (days_in_period / 7)
            elif expense.frequency == RegularExpenseFrequency.MONTHLY:
                regular_expenses += expense.amount * (days_in_period / 30)

        # Собираем данные по артикулам из report_data
        articles_data = {}
        for item in report_data:
            if not isinstance(item, dict):
                continue
                
            article = item.get("sa_name")
            if not article:
                if item.get("nm_id", 0):
                    # Ищем артикул по nm_id
                    for article2, item2 in articles_data.items():
                        if item2.get("nm_id") == item.get("nm_id"):
                            article = article2
                            break
            if not article:
                continue

            if article not in articles_data:
                articles_data[article] = {
                    "subject_name": item.get("subject_name", ""),
                    "orders": 0,
                    "sales": 0,
                    "returns": 0,
                    "cancellations": 0,
                    "sales_rub": 0,
                    "returns_rub": 0,
                    "commission": 0,
                    "logistics": 0,
                    "storage": 0,
                    "return_logistics": 0,
                    "acceptance": 0,
                    "corrections": 0,
                    "nm_id": item.get("nm_id", 0),
                    "deduction": 0
                }

            doc_type = item.get("doc_type_name", "")
            quantity = item.get("quantity", 0)
            price = item.get("retail_price_withdisc_rub", 0)
            retail_price = item.get("retail_price_withdisc_rub", 0)

            if "продажа" in doc_type.lower() or "sale" in doc_type.lower():
                articles_data[article]["sales"] += quantity
                articles_data[article]["sales_rub"] += price * quantity
            elif "возврат" in doc_type.lower() or "return" in doc_type.lower():
                articles_data[article]["returns"] += quantity
                articles_data[article]["returns_rub"] += price
            elif "отмена" in doc_type.lower() or "cancellation" in doc_type.lower():
                articles_data[article]["cancellations"] += quantity
            
            articles_data[article]["deduction"] += item.get('deduction', 0)
            
            # Правильный расчет комиссии WB
            # Комиссия = ppvz_sales_commission + ppvz_vw + ppvz_vw_nds
            commission = (
                item.get("ppvz_sales_commission", 0) +
                item.get("ppvz_vw", 0) +
                item.get("ppvz_vw_nds", 0)
            )
            articles_data[article]["commission"] += commission
            
            # Логистика = delivery_rub (только доставка)
            articles_data[article]["logistics"] += item.get("delivery_rub", 0)
            
            # Хранение = storage_fee
            articles_data[article]["storage"] += item.get("storage_fee", 0)
            
            # Приемка = acceptance (если есть)
            if "acceptance" in item:
                articles_data[article]["acceptance"] = articles_data[article].get("acceptance", 0) + item.get("acceptance", 0)
            
            # Корректировки = rebill_logistic_cost (если есть)
            if "rebill_logistic_cost" in item:
                articles_data[article]["corrections"] = articles_data[article].get("corrections", 0) + item.get("rebill_logistic_cost", 0)
            
            articles_data[article]["orders"] += quantity

        # Добавляем данные из заказов за период
        orders = (
            session.query(Order)
            .filter(Order.is_bouhght.is_(True))
            .filter(Order.date >= start_date)
            .filter(Order.date <= end_date)
            .filter(Order.isCancel.is_(False))
            .filter(Order.shop_id == shop_id)
            .all()
        )

        for order in orders:
            if order.supplierArticle not in articles_data:
                articles_data[order.supplierArticle] = {
                    "subject_name": order.supplierArticle,
                    "orders": 0,
                    "sales": 0,
                    "returns": 0,
                    "cancellations": 0,
                    "sales_rub": 0,
                    "returns_rub": 0,
                    "commission": 0,
                    "logistics": 0,
                    "storage": 0,
                    "return_logistics": 0,
                    "acceptance": 0,
                    "corrections": 0,
                    "nm_id": order.nmId,
                    "deduction": 0
                }
            articles_data[order.supplierArticle]["sales_rub"] += order.priceWithDisc
            articles_data[order.supplierArticle]["sales"] += 1
            articles_data[order.supplierArticle]["orders"] += 1
            articles_data[order.supplierArticle]["commission"] += order.priceWithDisc - order.forPay

        amount_articles = len(articles_data)

        # Обрабатываем общие удержания и хранение
        for item in report_data:
            if not isinstance(item, dict):
                continue
            if item.get("nm_id", 0) == 0:
                if item.get('bonus_type_name', '') == "Оказание услуг «ВБ.Продвижение»":
                    continue
                if item.get("ppvz_reward", 0):
                    for item2 in report_data:
                        if item2.get("srid") == item.get("srid"):
                            if item2.get("sa_name"):
                                articles_data[item2.get("sa_name")]["commission"] -= item.get("ppvz_reward")
                                break
                deduction = item.get("deduction", 0) / amount_articles if amount_articles > 0 else 0
                storage = item.get("storage_fee", 0) / amount_articles if amount_articles > 0 else 0
                for article, data in articles_data.items():
                    data["deduction"] += deduction
                    data["storage"] += storage

        regular_expenses_for_article = regular_expenses / amount_articles if amount_articles > 0 else 0

        # Получаем себестоимость товаров
        try:
            product_costs = (
                session.query(ProductCost).filter(ProductCost.shop_id == shop_id).all()
            )
            cost_map = {pc.article: pc.cost for pc in product_costs}
        except Exception as e:
            logger.error(f"Ошибка получения себестоимости: {e}")
            cost_map = {}

        # Получаем налоговую ставку
        try:
            tax_setting = (
                session.query(TaxSystemSetting)
                .filter(TaxSystemSetting.shop_id == shop_id)
                .first()
            )
            if tax_setting:
                if tax_setting.tax_system == TaxSystemType.USN_6:
                    tax_rate = 0.06
                elif tax_setting.tax_system == TaxSystemType.NO_TAX:
                    tax_rate = 0.0
                elif tax_setting.tax_system == TaxSystemType.CUSTOM:
                    tax_rate = tax_setting.custom_percent / 100 if tax_setting.custom_percent else 0.0
                else:
                    tax_rate = 0.0
        except Exception as e:
            logger.error(f"Ошибка получения налоговых настроек: {e}")
            tax_rate = 0.0

        # Заполняем данные в таблицу
        row_num = 2
        for article, data in articles_data.items():
            # Основные показатели
            revenue = data["sales_rub"] - data["returns_rub"]
            total_sales = data["sales"] - data["returns"]
            buyout_rate = (total_sales / data["orders"]) if data["orders"] else 0

            # Комиссии
            commission_percent = (data["commission"] / revenue) if revenue else 0

            # Логистика
            logistics_per_unit = data["logistics"] / total_sales if total_sales else 0
            logistics_percent = (data["logistics"] / revenue) if revenue else 0

            # Удержания
            total_deductions = (
                data["commission"]
                + data["logistics"]
                + data["return_logistics"]
                + data["storage"]
                + data["deduction"]
            )
            deductions_percent = (total_deductions / revenue) if revenue else 0

            # Налог
            tax = revenue * tax_rate

            # Себестоимость
            cost_per_item = cost_map.get(article, 0)
            total_cost = cost_per_item * total_sales

            # Прибыль
            profit_without_ads = (
                revenue - abs(total_cost) - abs(total_deductions) - abs(tax) - abs(regular_expenses_for_article)
            )

            # Рекламные расходы за период
            try:
                advertisement = sum(
                    i.amount
                    for i in session.query(Advertisement)
                    .filter(Advertisement.nmId == int(data["nm_id"]))
                    .filter(Advertisement.date >= start_date)
                    .filter(Advertisement.date <= end_date)
                    .all()
                )
            except Exception as e:
                logger.error(f"Ошибка получения рекламных расходов: {e}")
                advertisement = 0

            # Штрафы за период
            try:
                penalty = sum(
                    i.sum
                    for i in session.query(Penalty)
                    .filter(Penalty.nm_id == data["nm_id"])
                    .filter(Penalty.date >= start_date)
                    .filter(Penalty.date <= end_date)
                    .all()
                )
            except Exception as e:
                logger.error(f"Ошибка получения штрафов: {e}")
                penalty = 0

            profit_with_ads = (
                revenue
                - abs(total_cost)
                - abs(total_deductions)
                - abs(tax)
                - abs(advertisement)
                - abs(penalty)
                - abs(regular_expenses_for_article)
            )

            # Рентабельность
            profitability_cpm = (profit_without_ads / total_cost) * 100 if total_cost else 0
            profitability_sales = (profit_with_ads / revenue) * 100 if revenue else 0

            # Заполняем строку согласно новому порядку столбцов
            ws.cell(row=row_num, column=1, value=article)  # A - Артикул продавца
            ws.cell(row=row_num, column=2, value=data["subject_name"])  # B - Наименование
            ws.cell(row=row_num, column=3, value=profit_with_ads)  # C - Чистая прибыль
            # D - Не заполнять
            ws.cell(row=row_num, column=5, value=abs(data["orders"] * (data["sales_rub"] / data["sales"] if data["sales"] else 0)))  # E - Заказы руб
            ws.cell(row=row_num, column=6, value=abs(data["sales_rub"]))  # F - Выкупы руб
            ws.cell(row=row_num, column=7, value=abs(total_cost))  # G - Себестоимость
            ws.cell(row=row_num, column=8, value=abs(data["returns_rub"]))  # H - Возвраты (руб)
            ws.cell(row=row_num, column=9, value=abs(data["orders"]))  # I - Заказы (шт)
            ws.cell(row=row_num, column=10, value=abs(data["sales"]))  # J - Выкупы (шт)
            ws.cell(row=row_num, column=11, value=abs(data["returns"]))  # K - Возвраты (шт)
            # L - Не заполнять
            ws.cell(row=row_num, column=13, value=abs(data["storage"]))  # M - Хранение
            # N - Не заполнять
            ws.cell(row=row_num, column=15, value=data["commission"])  # O - Комиссия (руб)
            # P - Не заполнять
            ws.cell(row=row_num, column=17, value=abs(data["logistics"]))  # Q - Логистика (руб)
            # R - Не заполнять
            ws.cell(row=row_num, column=19, value=abs(logistics_per_unit))  # S - Логистика на ед
            ws.cell(row=row_num, column=20, value=advertisement)  # T - Реклама
            # U - Не заполнять
            ws.cell(row=row_num, column=22, value=penalty)  # V - Штрафы
            ws.cell(row=row_num, column=23, value=abs(data["deduction"]))  # W - Прочие удержания
            ws.cell(row=row_num, column=24, value=abs(tax))  # X - Налог

            row_num += 1

        # Итоговая строка
        last_row = ws.max_row + 1
        ws.cell(row=last_row, column=1, value="ИТОГО")
        for col in range(3, 25):  # Начиная с колонки "Заказы (шт)" до "Налог"
            col_letter = get_column_letter(col)
            if col in [5, 6, 11, 13, 16, 18, 20, 24]:  # Столбцы для средних значений
                ws.cell(row=last_row, column=col, value=f"=AVERAGE({col_letter}2:{col_letter}{last_row - 1})")
            else:  # Сумма по остальным столбцам
                ws.cell(row=last_row, column=col, value=f"=SUM({col_letter}2:{col_letter}{last_row - 1})")

        
        return wb
    except Exception as e:
        logger.error(f"Ошибка в generate_product_analytics_report: {e}")
        return None
    finally:
        session.close()

def apply_excel_formatting(ws):
    """Применяет форматирование к Excel-листу"""
    # Устанавливаем ширину столбцов
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # Форматирование чисел
    for row in ws.iter_rows(
        min_row=2, max_row=ws.max_row, min_col=3, max_col=ws.max_column
    ):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                # Проценты
                if cell.column_letter in ["E", "K", "M", "P", "T"]:
                    cell.number_format = "0.00%"
                elif (
                    cell.column >= 7
                    and cell.column <= 24
                    and cell.column not in [10, 11]
                ):
                    cell.number_format = "#,##0.00"
                else:
                    cell.number_format = "#,##0"

    # Границы
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            try:
                cell.border = thin_border
            except Exception as e:
                # Если не удается установить границу, пропускаем
                pass

    # Выравнивание заголовков
    for cell in ws[1]:
        try:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        except Exception as e:
            # Если не удается установить выравнивание, пропускаем
            pass

    # Фиксируем заголовки
    ws.freeze_panes = "A2"


async def back_to_analytics(callback: types.CallbackQuery, state: FSMContext):
    await analytics_callback(callback)

async def finances_handler(callback: types.CallbackQuery, state: FSMContext):
    text = "<b>Это раздел финансов</b>\n\nЗдесь Вы можете узнать свои главные показатели по своему бизнесу.\n\n▫️ Чистая прибыль\n▫️ Сроки окупаемости с учетом всех Ваших первоначальных затрат\n▫️ Рентабельность инвестиций покажет, насколько выгоден Ваш проект и как быстро он окупается"
    session = sessionmaker(bind=engine)()
    print(callback.from_user.id)
    user = session.query(User).filter(User.telegram_id == callback.from_user.id).first()
    session.close()

    if user:
        if user.subscription_end <= datetime.now():
            text += '\n\n⚠️ У вас закончилась подписка <b>JustProfit Premium</b>. Продлите её и Вам сразу же будут доступны все функции бота. \n\nПродлить подписку: "Главное меню" -> "Поддержка" -> "Подписка"'
            kb = InlineKeyboardMarkup()
            kb.add(InlineKeyboardButton("Меню", callback_data="main_menu"))
            await callback.message.delete()
            await callback.message.answer(text, reply_markup=kb)
            return
    kb = InlineKeyboardMarkup()
    kb.add(InlineKeyboardButton("Чистая прибыль", callback_data="an_1"))
    #kb.add(InlineKeyboardButton("ROS(Рентабльность продаж)", callback_data="an_2"))
    kb.add(InlineKeyboardButton("Срок окупаемости", callback_data="an_3"))
    kb.add(InlineKeyboardButton("ROI(Рентабельность вложений)", callback_data="an_4"))
    kb.add(InlineKeyboardButton("Меню", callback_data="main_menu"))
    await callback.message.delete()
    await callback.message.answer(text, reply_markup=kb)


async def pnl_callback(callback: types.CallbackQuery, state: FSMContext):
    # Проверяем выбран ли магазин
    async with state.proxy() as data:
        if "shop" not in data:
            await callback.answer("❌ Сначала выберите магазин", show_alert=True)
            return

    await callback.message.edit_text(
        "📊 <b>Расчёт прибыли и убытков (P&L)</b>\n\n" "Выберите период для расчета:",
    )


# Расчет показателей на основе отчета
async def calculate_metrics_from_report(full_report, shop_id, start_date, end_date, type_data="week", calculate_current_week=True):
    session = sessionmaker()(bind=engine)
    try:
        # Проверяем тип данных и извлекаем нужные части
        if isinstance(full_report, dict):
            report_data = full_report.get('finance', [])
            orders = full_report.get('orders', [])
            sales = full_report.get('sales', [])
        elif isinstance(full_report, list):
            # Если передан список, считаем что это finance данные
            report_data = full_report
            orders = []
            sales = []
        else:
            # Fallback
            report_data = []
            orders = []
            sales = []

        star = datetime.today() - timedelta(days=datetime.today().isoweekday())
        week_start = datetime(star.year, star.month, star.day + 1, 0, 0)
        print(start_date, end_date)
        
        # Используем логику как в pnl.py - создаем DataFrame
        df_orders = pd.DataFrame(orders)
        df_sales = pd.DataFrame(sales)
        df_fin = pd.DataFrame(report_data)
        
        # Определяем период
        DATE_FROM = start_date.strftime("%Y-%m-%d")
        DATE_TO = end_date.strftime("%Y-%m-%d")

        def determine_period_type(date_from, date_to):
            """Определяет тип периода и возвращает расширенный период для загрузки."""
            start_dt = dt.strptime(date_from, "%Y-%m-%d")
            end_dt = dt.strptime(date_to, "%Y-%m-%d")
            
            days_diff = (end_dt - start_dt).days
            
            if days_diff == 0:
                # День - загружаем неделю
                period_start = start_dt - timedelta(days=3)
                period_end = start_dt + timedelta(days=3)
                period_type = "день"
            elif days_diff <= 7:
                # Неделя - загружаем месяц
                period_start = start_dt - timedelta(days=15)
                period_end = end_dt + timedelta(days=15)
                period_type = "неделя"
            elif days_diff <= 31:
                # Месяц - загружаем квартал
                period_start = start_dt - timedelta(days=30)
                period_end = end_dt + timedelta(days=30)
                period_type = "месяц"
            else:
                # Год - загружаем полгода
                period_start = start_dt - timedelta(days=180)
                period_end = end_dt + timedelta(days=180)
                period_type = "год"
            
            return period_start.strftime("%Y-%m-%d"), period_end.strftime("%Y-%m-%d"), period_type

        # Определяем расширенный период
        extended_date_from, extended_date_to, period_type = determine_period_type(DATE_FROM, DATE_TO)
        
        # Ограничиваем максимальный период запроса
        start_dt = dt.strptime(extended_date_from, "%Y-%m-%d")
        end_dt = dt.strptime(extended_date_to, "%Y-%m-%d")
        
        # Максимальный период - 6 месяцев
        max_period_days = 180
        if (end_dt - start_dt).days > max_period_days:
            # Если период слишком большой, ограничиваем его
            if period_type == "год":
                start_dt = end_dt - timedelta(days=max_period_days)
            else:
                end_dt = start_dt + timedelta(days=max_period_days)
            extended_date_from = start_dt.strftime("%Y-%m-%d")
            extended_date_to = end_dt.strftime("%Y-%m-%d")
        
        # day‑столбец
        if not df_orders.empty:
            df_orders["day"] = pd.to_datetime(df_orders["date"]).dt.date
        else:
            df_orders["day"] = pd.to_datetime("2025-01-01").date()
            
        if not df_sales.empty:
            df_sales["day"] = pd.to_datetime(df_sales["date"]).dt.date
        else:
            df_sales["day"] = pd.to_datetime("2025-01-01").date()
        
        # Проверяем наличие поля rr_dt в финансовых данных
        if not df_fin.empty:
            if "rr_dt" in df_fin.columns:
                df_fin["day"] = pd.to_datetime(df_fin["rr_dt"]).dt.date
            elif "date" in df_fin.columns:
                df_fin["day"] = pd.to_datetime(df_fin["date"]).dt.date
            elif "sale_dt" in df_fin.columns:
                df_fin["day"] = pd.to_datetime(df_fin["sale_dt"]).dt.date
            else:
                # Если нет поля даты, создаем пустой столбец day
                df_fin["day"] = pd.to_datetime("2025-01-01").date()
        else:
            df_fin["day"] = pd.to_datetime("2025-01-01").date()

        # 4.1 Orders
        if not df_orders.empty:
            grp_orders = (
                df_orders.groupby("day")
                        .agg(
                            order_sum = ("priceWithDisc", "sum"),
                            ordered_qty = ("srid", "size"),
                            orders_cnt = ("srid", "nunique")
                        )
            )
        else:
            # Создаем пустой DataFrame с нужными столбцами
            grp_orders = pd.DataFrame(columns=['order_sum', 'ordered_qty', 'orders_cnt'])

        # 4.2 Sales (только продажи, не возвраты)
        if not df_sales.empty:
            sales_mask = df_sales["saleID"].str.startswith("S", na=False)
            grp_sales = (
                df_sales[sales_mask]
                .groupby("day")
                .agg(
                    sales_sum = ("priceWithDisc", "sum"),
                    sold_qty = ("saleID", "size"),
                    for_pay = ("forPay", "sum")
                )
            )
        else:
            # Создаем пустой DataFrame с нужными столбцами
            grp_sales = pd.DataFrame(columns=['sales_sum', 'sold_qty', 'for_pay'])

        # 4.3 Finance — logistica = только delivery_rub
        if not df_fin.empty:
            grp_fin = (
                df_fin.groupby("day")
                    .agg(
                        delivery_cost = ("delivery_rub", "sum"),
                        storage = ("storage_fee", "sum"),
                        penalty = ("penalty", "sum"),
                        acceptance = ("acceptance", "sum"),
                        pay_for_goods = ("ppvz_for_pay", "sum")
                    )
                    .apply(pd.to_numeric, errors="coerce")
            )
        else:
            # Создаем пустой DataFrame с нужными столбцами
            grp_fin = pd.DataFrame(columns=['delivery_cost', 'storage', 'penalty', 'acceptance', 'pay_for_goods'])

        # ────────────────────────────────────────────────────────────────────
        # 5. Объединяем и рассчитываем
        # ────────────────────────────────────────────────────────────────────
        all_daily = (
            grp_orders
            .join(grp_sales, how="outer")
            .join(grp_fin, how="outer")
            .fillna(0)
            .infer_objects(copy=False)
        )

        all_daily["buyout_pct"] = (
            all_daily["sold_qty"] / all_daily["ordered_qty"]
        ).fillna(0)

        all_daily["total_to_pay"] = (
            all_daily["pay_for_goods"]
            - all_daily[["delivery_cost", "storage", "penalty", "acceptance"]].sum(axis=1)
        )

        # Фильтруем данные для текущего периода
        current_start = start_date.date()
        current_end = end_date.date()
        daily = all_daily[(all_daily.index >= current_start) & (all_daily.index <= current_end)]
        
        # Отладочная информация
        print(f"Период: {current_start} - {current_end}")
        print(f"Всего записей в all_daily: {len(all_daily)}")
        print(f"Записей в daily после фильтрации: {len(daily)}")
        if not daily.empty:
            print(f"Индексы daily: {daily.index.tolist()}")

        # Получаем рекламу и штрафы из базы данных
        advert = sum(
            i.amount for i in session.query(Advertisement)
            .filter(Advertisement.shop_id == shop_id)
            .filter(Advertisement.date >= start_date)
            .filter(Advertisement.date <= end_date)
        )

        stops = sum(
            i.amount for i in session.query(Penalty)
            .filter(Penalty.shop_id == shop_id)
            .filter(Penalty.date >= start_date)
            .filter(Penalty.date <= end_date)
        )

        # Получаем настройки налоговой системы
        tax_setting = session.query(TaxSystemSetting).filter(TaxSystemSetting.shop_id == shop_id).first()
        
        # Рассчитываем налог
        revenue = daily["sales_sum"].sum()
        if tax_setting:
            if tax_setting.tax_system == TaxSystemType.USN_6:
                tax_rate = 0.06
                print("tax_rate = ", tax_rate)
            elif tax_setting.tax_system == TaxSystemType.NO_TAX:
                tax_rate = 0.0
                print("tax_rate = ", tax_rate)
            elif tax_setting.tax_system == TaxSystemType.CUSTOM:
                tax_rate = tax_setting.custom_percent / 100
                print("tax_rate = ", tax_rate)
            else:
                tax_rate = 0.0
                print("пошло в else")
        else:
            tax_rate = 0.0
            
        tax = revenue * tax_rate

        # Себестоимость (как в pnl.py)
        cost_of_goods = 0
        articles = {}
        
        # Собираем артикулы для расчета себестоимости
        for item in report_data:
            if not isinstance(item, dict):
                continue
            article = item.get("nm_id")
            quantity = item.get("quantity", 0)
            if article and quantity:
                if article not in articles:
                    articles[article] = 0
                articles[article] += quantity

        # Рассчитываем себестоимость
        for article, quantity in articles.items():
            print(article, quantity)
            try:
                supp_article = session.query(Order).filter(Order.nmId == int(article)).first().supplierArticle
                product_cost = (
                    session.query(ProductCost)
                    .filter(ProductCost.shop_id == shop_id, ProductCost.article == supp_article)
                    .first()
                )
                if product_cost:
                    cost_of_goods += product_cost.cost * quantity
            except:
                pass

        # Регулярные затраты за период
        regular_expenses = 0
        days_in_period = (end_date - start_date).days + 1
        # print(start_date, end_date, days_in_period)
        for expense in session.query(RegularExpense).filter(RegularExpense.shop_id == shop_id):
            if type_data == "week":
                if expense.frequency == RegularExpenseFrequency.WEEKLY:
                    regular_expenses += expense.amount * 1
                if expense.frequency == RegularExpenseFrequency.DAILY:
                    regular_expenses += expense.amount * 7
            if type_data == "month":
                if expense.frequency == RegularExpenseFrequency.DAILY:
                    regular_expenses += expense.amount * 30
                if expense.frequency == RegularExpenseFrequency.WEEKLY:
                    regular_expenses += expense.amount * 4
                if expense.frequency == RegularExpenseFrequency.MONTHLY:
                    regular_expenses += expense.amount * 1
            if type_data == "year":
                if expense.frequency == RegularExpenseFrequency.DAILY:
                    regular_expenses += expense.amount * 365
                if expense.frequency == RegularExpenseFrequency.WEEKLY:
                    regular_expenses += expense.amount * 52
                if expense.frequency == RegularExpenseFrequency.MONTHLY:
                    regular_expenses += expense.amount * 12

        # Чистая прибыль (как в pnl.py)
        commission = daily["pay_for_goods"].sum()
        logistics = daily["delivery_cost"].sum()
        storage_fee = daily["storage"].sum()
        deduction = daily["penalty"].sum() + stops
        
        net_profit = revenue - commission - logistics - storage_fee - tax - cost_of_goods - regular_expenses - deduction - advert

        # Рентабельность (как в pnl.py)
        profitability = (net_profit / revenue) * 100 if revenue > 0 else 0

        # Рекламные затраты

        # Разовые затраты (инвестиционные)
        one_time_expenses = (
            session.query(OneTimeExpense)
            .filter(OneTimeExpense.shop_id == shop_id)
            .all()
        )
        total_one_time = sum(expense.amount for expense in one_time_expenses)

        # Срок окупаемости будет рассчитываться в блоке годовой доходности
        payback_period = "будет рассчитан в годовом отчете"

        # ROI
        roi = "не определен"
        if total_one_time > 0:
            roi_value = (net_profit / total_one_time) * 100
            roi = f"{roi_value:.1f}%"
            if roi_value > 100:
                roi += " ✅ Поздравляем, вы окупили вложения!"

        try:
            ros_value = (net_profit / revenue) * 100
        except:
            ros_value = 0
        return {
            "revenue": revenue,
            "commission": commission,
            "logistics": logistics,
            "storage": storage_fee,
            "cost_of_goods": cost_of_goods,
            "tax": tax,
            "regular_expenses": regular_expenses,
            "net_profit": net_profit,
            "profitability": profitability,
            "payback_period": payback_period,
            "roi": roi,
            "total_one_time": total_one_time,
            "advert": advert,
            "stops": stops,
            "deduction": deduction,
            "ros": ros_value,
        }
    finally:
        session.close()


async def select_anal_period_callback(callback: types.CallbackQuery, state: FSMContext):
    period_type = callback.data.split("_")[1]  # day, week, month, year, custom
    
    # Получаем данные состояния
    data = await state.get_data()
    shop_data = data.get("shop")
    
    if not shop_data:
        await callback.answer("❌ Сначала выберите магазин", show_alert=True)
        return
        
    shop_id = shop_data["id"]
    shop_name = shop_data.get("name") or f"Магазин {shop_id}"
    api_token = shop_data["api_token"]
    an_type = data.get("an_type", "an_1")  # Используем значение по умолчанию
    
    # Проверяем и исправляем an_type если нужно
    if not an_type or not an_type.startswith("an_"):
        an_type = "an_1"
        await state.update_data(an_type=an_type)
    
    await callback.message.delete()
    message = await callback.message.answer(
        text="Производим расчёт данных, пожалуйста, подождите\n\n"
             "‼️ Важно: <u>необходимо подождать до 1 минуты для полного завершения расчёта "
             "(Ограничение Wildberries на количество запросов)</u>"
    )

    now = datetime.now()

    # Определяем даты периода
    if period_type == "custom" or period_type.startswith("custom_"):
        # Проверяем, есть ли кастомные даты
        if data.get("custom_period") and data.get("custom_start_date") and data.get("custom_end_date"):
            current_start = data["custom_start_date"]
            current_end = data["custom_end_date"]
            type_datalol = "custom"  # Всегда используем "custom" для кастомного периода
            period_name = f"{current_start.strftime('%d.%m.%Y')}-{current_end.strftime('%d.%m.%Y')}"
        else:
            # Фоллбек на текущий день, если кастомные даты не заданы
            current_start = datetime(now.year, now.month, now.day)
            current_end = now
            type_datalol = "day"
            period_name = f"{current_start.strftime('%d.%m')}-{now.strftime('%d.%m')}"
    elif period_type == "week":
        start_week = now - timedelta(days=now.isoweekday() - 1)
        current_start = datetime(start_week.year, start_week.month, start_week.day)
        current_end = now
        type_datalol = "week"
        period_name = f"{current_start.strftime('%d.%m')}-{now.strftime('%d.%m')}"
    elif period_type == "month":
        current_start = datetime(now.year, now.month, 1)
        current_end = now
        type_datalol = "month"
        period_name = f"{current_start.strftime('%d.%m')}-{now.strftime('%d.%m')}"
    elif period_type == "year":
        current_start = datetime(now.year, 1, 1)
        current_end = now
        type_datalol = "year"
        period_name = f"{current_start.strftime('%d.%m')}-{now.strftime('%d.%m')}"
    else:
        # По умолчанию - неделя
        start_week = now - timedelta(days=now.isoweekday() - 1)
        current_start = datetime(start_week.year, start_week.month, start_week.day)
        current_end = now
        type_datalol = "week"
        period_name = f"{current_start.strftime('%d.%m')}-{now.strftime('%d.%m')}"

    # Получаем данные напрямую из API (с кэшированием)
    loop = asyncio.get_event_loop()
    full_report = await loop.run_in_executor(
        None,
        fetch_report_detail_by_period_cached,
        api_token,
        current_start,
        current_end,
        shop_id
    )
    
    # Проверяем структуру данных
    if isinstance(full_report, dict):
        current_report = full_report.get('finance', [])
    else:
        # Если full_report не словарь, считаем что это finance данные
        current_report = full_report if isinstance(full_report, list) else []
    # Если нет данных и период не неделя — предупреждаем
    if not current_report and period_type != "week":
        await callback.answer(
            "❌ Не удалось получить данные за текущий период, подождите около 1-2 минуты и попробуйте снова",
            show_alert=True,
        )
        return


    # Получаем type_data из an_type с проверкой
    try:
        if an_type and an_type.startswith("an_"):
            type_data = int(an_type.split("_")[1])
        else:
            # Если an_type не в правильном формате, используем значение по умолчанию
            type_data = 1
            logger.warning(f"an_type имеет неожиданный формат: {an_type}, используем type_data=1")
    except (ValueError, IndexError, AttributeError) as e:
        # Если не удается распарсить, используем значение по умолчанию
        type_data = 1
        logger.warning(f"Ошибка парсинга an_type '{an_type}': {e}, используем type_data=1")

               
    
        
    # Убираем переопределение дат для type_data == 3, чтобы кастомные даты работали для всех типов
    # if type_data == 3:
    #     current_start = datetime(now.year, now.month, 1)
    #     current_end = now
    #     type_datalol = "month"
    #     # Получаем данные за месяц напрямую из API
    #     loop = asyncio.get_event_loop()
    #     current_report = await loop.run_in_executor(
    #         None,
    #         fetch_report_detail_by_period,
    #         api_token,
    #         current_start,
    #         current_end
    #     )




    if not current_report and period_type != "week":
        await callback.answer(
            "❌ Не удалось получить данные за текущий период, подождите около 1-2 минуты и попробуйте снова",
            show_alert=True,
        )
        return
    # print(an_type)
    logger.info(f"Create AN report. AN-type: {an_type}")
    
    # Получаем данные напрямую из API вместо кэша
    loop = asyncio.get_event_loop()
    
    # Рассчитываем показатели для выбранного периода
    current_metrics = await calculate_metrics_from_report(
        full_report, shop_id, current_start, current_end, type_datalol
    )

    # Рассчитываем динамику
    # revenue_change = current_metrics["revenue"] - (previous_metrics["revenue"] if previous_metrics else 0)
    # profit_change = current_metrics["net_profit"] - (previous_metrics["net_profit"] if previous_metrics else 0)

    # revenue_indicator = "🟢▲" if revenue_change >= 0 else "🔴▼"
    # profit_indicator = "🔴▼" if profit_change >= 0 else "🔴▼"
    


    # Форматируем отчет
    text = ""

    async with state.proxy() as data:
        an_type = data["an_type"]
    
    # Создаем клавиатуру
    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("🔙 Назад", callback_data="main_menu"))
    
    if True:
        # Блок an_1
        last_metrics = None  # Убираем сравнение с предыдущим периодом
        destanation = "Нет данных для сравнения"
        text = (
            f"Период: <b>({period_name})</b>\n\n"
            "<u>Основные показатели:</u>\n"
            f"▫️Выручка: {current_metrics['revenue']:.2f} руб.\n"
            f"▫️Комиссии: {current_metrics['commission']:.2f} руб. <b>{current_metrics['commission']/current_metrics['revenue']*100:.1f}%</b>\n"
            f"▫️Логистика: {current_metrics['logistics']:.2f} руб. <b>{current_metrics['logistics']/current_metrics['revenue']*100:.1f}%</b>\n"
            f"▫️Хранение: {current_metrics['storage']:.2f} руб. <b>{current_metrics['storage']/current_metrics['revenue']*100:.1f}%</b>\n"
            f"▫️Себестоимость: {current_metrics['cost_of_goods']:.2f} руб. <b>{current_metrics['cost_of_goods']/current_metrics['revenue']*100:.1f}%</b>\n"
            f"▫️Налог: {current_metrics['tax']:.2f} руб. <b>{current_metrics['tax']/current_metrics['revenue']*100:.1f}%</b>\n"
            f"▫️Рекламные затраты: {current_metrics['advert']} руб. <b>{current_metrics['advert']/current_metrics['revenue']*100:.1f}%</b>\n"
            f"▫️Прочие удержания: {current_metrics['deduction']} руб. <b>{current_metrics['deduction']/current_metrics['revenue']*100:.1f}%</b>\n"
            f"▫️Штрафы: {current_metrics['stops']} руб. <b>{current_metrics['stops']/current_metrics['revenue']*100:.1f}%</b>\n\n"
            f"〽️ Чистая прибыль: {current_metrics['net_profit']:.2f} руб. ({destanation}) <b>{current_metrics['net_profit']/current_metrics['revenue']*100:.1f}%</b>\n\n"
        )

        text += f"\n<i>Примечание: расчеты основаны на данных WB API</i>"

        # Генерируем Excel-файлы
        pnl_excel_path = None
        product_excel_path = None
        
        try:
            # Генерируем PNL отчет
            wb_pnl = await generate_pnl_excel_report(
                shop_id,
                api_token,
                current_start,
                current_end,
                shop_name,
                full_report
            )
            if wb_pnl:
                pnl_excel_path = f"PNL.xlsx"
                wb_pnl.save(pnl_excel_path)
        except Exception as e:
            logger.error(f"Ошибка при генерации PNL Excel-отчёта: {e}")
            pnl_excel_path = None

        try:
            # Генерируем отчет по товарам
            wb_product = await generate_product_analytics_report(
                api_token,
                shop_id,
                current_start,
                current_end,
                full_data=full_report  # Передаем full_data
            )
            if wb_product:
                product_excel_path = f"отчёт по товарам .xlsx"
                wb_product.save(product_excel_path)
        except Exception as e:
            logger.error(f"Ошибка при генерации товарного Excel-отчёта: {e}")
            product_excel_path = None

        # Отправляем сообщение с текстом
        await message.edit_text(text, reply_markup=keyboard)
        
        # Отправляем Excel-файлы отдельно
        if pnl_excel_path and os.path.exists(pnl_excel_path):
            await message.answer_document(
                InputFile(pnl_excel_path)
            )
            os.remove(pnl_excel_path)
            
        if product_excel_path and os.path.exists(product_excel_path):
            await message.answer_document(
                InputFile(product_excel_path)
            )
            os.remove(product_excel_path)
    
    # Проверяем, что text не пустой
    if not text:
        text = "❌ Ошибка: не удалось сформировать отчет"
    # await state.finish()
    # kb.add(InlineKeyboardButton("Чистая прибыль", callback_data="an_1"))
    # kb.add(InlineKeyboardButton("ROS(Рентабльность продаж)", callback_data="an_2"))
    # kb.add(InlineKeyboardButton("Срок окупаемости", callback_data="an_3"))
    # kb.add(InlineKeyboardButton("ROI(Рентабельность вложений)", callback_data="an_4"))



async def anal_callback(callback: types.CallbackQuery, state: FSMContext):
    # Проверяем выбран ли магазин

    async with state.proxy() as data:
        if "shop" not in data:
            await callback.answer("❌ Сначала выберите магазин", show_alert=True)
            return
        print(callback.data)
        data["an_type"] = callback.data
    
    # Проверяем формат callback.data
    try:
        type_data = int(callback.data.split("_")[1])
    except (ValueError, IndexError):
        # Если формат неправильный, используем значение по умолчанию
        type_data = 1
        callback.data = "an_1"
        data["an_type"] = "an_1"
    
    # print(type_data)
    text = ""
    keyboard = None
    
    if type_data == 1:
        text = "Чистой прибыли"
        keyboard = period_keyboard(type_data)
    elif type_data == 2:
        text = "ROS (Рентабельности продаж)"
        keyboard = period_keyboard2(type_data)
    elif type_data == 3:
        text = "Срока окупаемости"
        return await select_anal_period_callback(callback, state)
    elif type_data == 4:
        text = "Рентабельность вложений"
        return await select_anal_period_callback(callback, state)

    # Проверяем, что text и keyboard установлены
    if not text or not keyboard:
        await callback.answer("❌ Ошибка в настройке меню", show_alert=True)
        return

    # print(text)
    await callback.message.delete()
    await callback.message.answer(
        f" <b>Расчёт {text}</b>\n\n" "Выберите период для расчета:",
        reply_markup=keyboard,
    )

async def custom_period_callback(callback: types.CallbackQuery, state: FSMContext):
    """Обработчик для кнопки 'Выбранный период' - сразу показывает календарь"""
    type_data = callback.data.split("_")[-1]
    
    async with state.proxy() as data:
        data["an_type"] = f"an_{type_data}"
        data["period_size"] = "custom"  # Устанавливаем кастомный период
        # Очищаем предыдущие даты
        data.pop("first_date", None)
        data.pop("second_date", None)
    
    # Удаляем предыдущее сообщение
    try:
        await callback.message.delete()
    except:
        pass
    
    # Сразу показываем календарь для выбора дат
    await show_calendar_for_custom_period(callback, state)

async def custom_period_size_callback(callback: types.CallbackQuery, state: FSMContext):
    """Обработчик выбора размера периода - показывает календарь"""
    parts = callback.data.split("_")
    period_size = parts[1]  # day, week, month
    type_data = parts[2]
    
    async with state.proxy() as data:
        data["period_size"] = period_size
        data["an_type"] = f"an_{type_data}"
    
    # Показываем календарь для выбора даты
    await show_calendar(callback, state, period_size)

async def show_calendar(callback: types.CallbackQuery, state: FSMContext, period_size):
    now = datetime.now()
    async with state.proxy() as data:
        # Только если еще не установлено!

        current_month = data.get("calendar_month", now.month)
        current_year = data.get("calendar_year", now.year)

        # Если это первый запуск, сохрани в state
        data["calendar_month"] = current_month
        data["calendar_year"] = current_year

    await show_calendar_for_month(callback, state, period_size, current_month, current_year)

async def custom_period_back_callback(callback: types.CallbackQuery, state: FSMContext):
    """Обработчик кнопки 'Назад' в календаре"""
    async with state.proxy() as data:
        an_type = data.get("an_type", "an_1")
        type_data = an_type.split("_")[1]
    
    # Создаем фейковый callback с правильным форматом
    callback.data = an_type
    
    # Возвращаемся к выбору типа аналитики
    await anal_callback(callback, state)

async def select_date_callback(callback: types.CallbackQuery, state: FSMContext):
    """Обработчик выбора даты из календаря"""
    parts = callback.data.split("_")
    period_size = parts[2]  # day, week, month
    date_str = parts[3]     # YYYY-MM-DD
    
    selected_date = datetime.strptime(date_str, "%Y-%m-%d")
    
    async with state.proxy() as data:
        data["selected_date"] = selected_date
        data["period_size"] = period_size
    
    # Сразу подтверждаем выбор без промежуточного меню
    await confirm_custom_callback(callback, state)

async def show_date_confirmation(callback: types.CallbackQuery, state: FSMContext, selected_date, period_size):
    """Показывает подтверждение выбранной даты"""
    period_text = {"day": "день", "week": "неделю", "month": "месяц"}[period_size]
    
    # Рассчитываем период в зависимости от выбора
    if period_size == "day":
        start_date = selected_date
        end_date = selected_date
        period_display = selected_date.strftime("%d.%m.%Y")
    elif period_size == "week":
        # Находим понедельник недели
        start_date = selected_date - timedelta(days=selected_date.isoweekday() - 1)
        end_date = start_date + timedelta(days=6)
        period_display = f"{start_date.strftime('%d.%m')}-{end_date.strftime('%d.%m.%Y')}"
    else:  # month
        start_date = selected_date.replace(day=1)
        end_date = (start_date.replace(month=start_date.month % 12 + 1, day=1) - timedelta(days=1))
        period_display = f"{start_date.strftime('%d.%m')}-{end_date.strftime('%d.%m.%Y')}"
    
    async with state.proxy() as data:
        data["custom_start_date"] = start_date
        data["custom_end_date"] = end_date
    
    keyboard = InlineKeyboardMarkup(row_width=1)
    keyboard.add(
        InlineKeyboardButton("✅ Подтвердить", callback_data=f"confirm_custom_{period_size}"),
        InlineKeyboardButton("🔄 Выбрать другую дату", callback_data=f"custom_{period_size}_{data['an_type'].split('_')[1]}"),
        InlineKeyboardButton("🔙 Назад", callback_data="custom_period_back")
    )
    
    try:
        await callback.message.edit_text(
            f"📅 <b>Подтверждение выбора</b>\n\n"
            f"Выбранный {period_text}: <b>{period_display}</b>\n\n"
            f"Нажмите 'Подтвердить' для расчета:",
            reply_markup=keyboard
        )
    except MessageNotModified:
        # Если сообщение не изменилось, просто отвечаем на callback
        await callback.answer()
    except Exception as e:
        # Для других ошибок логируем и отвечаем на callback
        print(f"Error in show_date_confirmation: {e}")
        await callback.answer()


async def confirm_custom_callback(callback: types.CallbackQuery, state: FSMContext):
    """Обработчик подтверждения кастомного периода"""
    # Получаем данные состояния
    data = await state.get_data()
    selected_date = data.get("selected_date")
    period_size = data.get("period_size")
    an_type = data["an_type"]
    report_type = data.get("report_type")  # По умолчанию product_analytics
    
    # Рассчитываем период в зависимости от выбора
    if period_size == "day":
        start_date = selected_date
        end_date = selected_date
    elif period_size == "week":
        # Находим понедельник недели
        start_date = selected_date - timedelta(days=selected_date.isoweekday() - 1)
        end_date = start_date + timedelta(days=6)
    else:  # month
        start_date = selected_date.replace(day=1)
        end_date = (start_date.replace(month=start_date.month % 12 + 1, day=1) - timedelta(days=1))
    
    # Сохраняем кастомные даты в состоянии
    await state.update_data(
        custom_period=True,
        custom_start_date=start_date,
        custom_end_date=end_date
    )
    print("report_type", report_type)

    # Получаем данные магазина из состояния
    shop_data = data.get("shop")
    if not shop_data:
        await callback.answer("❌ Магазин не выбран. Сначала выберите магазин.", show_alert=True)
        return
        
    shop_id = shop_data["id"]
    api_token = shop_data["api_token"]
    
    # Вызываем основную функцию расчета
    if report_type == "product_analytics":
        await product_analytics_callback(callback, state, start_date, end_date)
    else:
        await select_anal_period_callback(callback, state)    

async def calendar_navigation_callback(callback: types.CallbackQuery, state: FSMContext):
    print(f"calendar_navigation_callback triggered: {callback.data}")
    parts = callback.data.split("_")
    action = parts[0]  # prev or next
    period_size = parts[2]  # day, week, month

    data = await state.get_data()
    current_month = data.get("calendar_month", datetime.now().month)
    current_year = data.get("calendar_year", datetime.now().year)

    # ⬅️➡️ изменяем месяц
    if action == "prev":
        if current_month == 1:
            current_month = 12
            current_year -= 1
        else:
            current_month -= 1
    elif action == "next":
        if current_month == 12:
            current_month = 1
            current_year += 1
        else:
            current_month += 1

    # 💾 Сохраняем обратно в FSM
    await state.update_data(calendar_month=current_month, calendar_year=current_year)

    try:
        await show_calendar_for_month(callback, state, period_size, current_month, current_year)
    except MessageNotModified:
        await callback.answer()
    except Exception as e:
        print(f"Error in calendar_navigation_callback: {e}")
        await callback.answer()

async def show_calendar_for_month(callback: types.CallbackQuery, state: FSMContext, period_size, month, year):
    """Показывает календарь для конкретного месяца"""
    print(f"Show calendar for: {month}.{year}")
    # Создаем календарь на указанный месяц
    keyboard = InlineKeyboardMarkup(row_width=7)
    
    # Заголовок месяца
    month_date = datetime(year, month, 1)
    month_name = month_date.strftime("%B %Y")
    keyboard.add(InlineKeyboardButton(f"📅 {month_name}", callback_data="ignore"))
    
    # Дни недели
    weekdays = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    keyboard.row(*[InlineKeyboardButton(day, callback_data="ignore") for day in weekdays])
    
    # Получаем первый день месяца и количество дней
    first_day = datetime(year, month, 1)
    if month == 12:
        next_month = datetime(year + 1, 1, 1)
    else:
        next_month = datetime(year, month + 1, 1)
    days_in_month = (next_month - timedelta(days=1)).day
    
    # Определяем день недели для первого дня (1 = понедельник)
    first_weekday = first_day.isoweekday()
    
    # Добавляем пустые ячейки в начале
    row = []
    for _ in range(first_weekday - 1):
        row.append(InlineKeyboardButton(" ", callback_data="ignore"))
    
    # Добавляем дни месяца
    now = datetime.now()
    for day in range(1, days_in_month + 1):
        date_str = f"{year}-{month:02d}-{day:02d}"
        callback_data = f"select_date_{period_size}_{date_str}"

        if day == now.day and month == now.month and year == now.year:
            row.append(InlineKeyboardButton(f"•{day}•", callback_data=callback_data))
        else:
            row.append(InlineKeyboardButton(str(day), callback_data=callback_data))

        if len(row) == 7:
            keyboard.row(*row)
            row = []

    # ДОБАВЛЯЕМ ОСТАВШИЕСЯ ТОЛЬКО ПОСЛЕ ЦИКЛА
    if row:
        while len(row) < 7:
            row.append(InlineKeyboardButton(" ", callback_data="ignore"))
        keyboard.row(*row)

    
    # Кнопки навигации (без ограничений по году)
    nav_row = [
        InlineKeyboardButton("◀️", callback_data=f"prev_month_{period_size}"),
        InlineKeyboardButton("🔙 Назад", callback_data="main_menu"),
        InlineKeyboardButton("▶️", callback_data=f"next_month_{period_size}")
    ]
    keyboard.row(*nav_row)
    
    period_text = {"day": "день", "week": "неделю", "month": "месяц"}[period_size]
    
    try:
        await callback.message.edit_text(
            f"📅 <b>Выберите {period_text} для расчета</b>\n\n"
            f"Нажмите на дату, чтобы выбрать {period_text}:",
            reply_markup=keyboard
        )
    except MessageNotModified:
        # Если сообщение не изменилось, просто отвечаем на callback
        await callback.answer()
    except Exception as e:
        # Для других ошибок логируем и отвечаем на callback
        print(f"Error in show_calendar_for_month: {e}")
        await callback.answer() 

async def ignore_callback(callback: types.CallbackQuery, state: FSMContext):
    """Обработчик для игнорируемых кнопок (заголовки, дни недели)"""
    await callback.answer()    

def register_analytics_handlers(dp: Dispatcher):
    #Обработчик под календарь на эксель отчёт
    dp.register_callback_query_handler(
        start_analytics_report, 
        lambda c: c.data == "start_analytics_report", 
        state="*"
    )
    dp.register_callback_query_handler(analytics_callback, text="analytics", state="*")
    dp.register_callback_query_handler(
        profitability_estimation_callback, text="profitability_estimation", state="*"
    )
    dp.register_callback_query_handler(
        top5_products_callback, text="top5_products", state="*"
    )
    dp.register_callback_query_handler(
        what_if_simulator_callback, text="what_if_simulator", state="*"
    )
    dp.register_callback_query_handler(
        product_analytics_callback, text="product_analytics", state="*"
    )
    
    dp.register_callback_query_handler(finances_handler, text="finances", state="*")
    # Пагинация и выбор артикула
    dp.register_callback_query_handler(
        handle_articles_pagination,
        lambda c: c.data in ["prev_articles_page", "next_articles_page"],
        state=AnalyticsStates.waiting_for_article,
    )
    dp.register_callback_query_handler(
        select_article_callback,
        lambda c: c.data.startswith("select_article_"),
        state=AnalyticsStates.waiting_for_article,
    )
    dp.register_callback_query_handler(
        anal_callback, lambda c: c.data.startswith("an_"), state="*"
    )

    dp.register_callback_query_handler(
    ignore_callback, 
    lambda c: c.data == "ignore", 
    state="*"
    )
    # Новые обработчики для кастомного периода (должны быть выше)
    dp.register_callback_query_handler(
        select_custom_date_callback, 
        lambda c: c.data.startswith("select_custom_date_any_"), 
        state="*"
    )
    
    dp.register_callback_query_handler(
        confirm_custom_period_callback, 
        lambda c: c.data == "confirm_custom_period", 
        state="*"
    )
    
    dp.register_callback_query_handler(
        restart_custom_period_callback, 
        lambda c: c.data == "restart_custom_period", 
        state="*"
    )
    
    dp.register_callback_query_handler(
        custom_period_navigation_callback, 
        lambda c: c.data.startswith("prev_month_custom_any") or c.data.startswith("next_month_custom_any"), 
        state="*"
    )
    
    # Обработчики для календаря
    dp.register_callback_query_handler(
        calendar_navigation_callback, 
        lambda c: (c.data.startswith("prev_month_") or c.data.startswith("next_month_")) and not c.data.startswith("prev_month_custom_") and not c.data.startswith("next_month_custom_"), 
        state="*"
    )
    
    dp.register_callback_query_handler(
        custom_period_back_callback, 
        lambda c: c.data == "custom_period_back", 
        state="*"
    )

        # Обработчик для "Выбранный период"
    dp.register_callback_query_handler(
        custom_period_callback, 
        lambda c: c.data.startswith("custom_period_"), 
        state="*"
    )

    dp.register_callback_query_handler(
        custom_period_size_callback, 
        lambda c: c.data.startswith("custom_") and not c.data.startswith("custom_period_"), 
        state="*"
    )
    
    dp.register_callback_query_handler(
        select_date_callback, 
        lambda c: c.data.startswith("select_date_"), 
        state="*"
    )
    
    dp.register_callback_query_handler(
        confirm_custom_callback, 
        lambda c: c.data.startswith("confirm_custom_"), 
        state="*"
    )
        
    dp.register_callback_query_handler(
        select_anal_period_callback, lambda c: c.data.startswith("anperiod_"), state="*"
    )
    # Ввод для симулятора
    dp.register_message_handler(
        process_price_and_cost, state=AnalyticsStates.waiting_for_price_and_cost
    )

    # Возврат в меню
    dp.register_callback_query_handler(
        back_to_analytics, text="back_to_analytics", state="*"
    )

async def show_calendar_for_custom_period(callback: types.CallbackQuery, state: FSMContext):
    """Показывает календарь для выбора кастомного периода (две даты)"""
    now = datetime.now()
    async with state.proxy() as data:
        current_month = data.get("calendar_month", now.month)
        current_year = data.get("calendar_year", now.year)
        
        # Сохраняем в state
        data["calendar_month"] = current_month
        data["calendar_year"] = current_year
    
    # Удаляем предыдущее сообщение, если оно есть
    try:
        await callback.message.delete()
    except:
        pass
    
    # Создаем новое сообщение с календарем
    await show_calendar_for_month_custom(callback, state, current_month, current_year, "any")

async def show_calendar_for_month_custom(callback: types.CallbackQuery, state: FSMContext, month, year, step):
    """Показывает календарь для выбора кастомного периода"""
    print(f"Show custom calendar for: {month}.{year}")
    
    # Создаем календарь на указанный месяц
    keyboard = InlineKeyboardMarkup(row_width=7)
    
    # Заголовок месяца
    month_date = datetime(year, month, 1)
    month_name = month_date.strftime("%B %Y")
    keyboard.add(InlineKeyboardButton(f"📅 {month_name}", callback_data="ignore"))
    
    # Дни недели
    weekdays = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    keyboard.row(*[InlineKeyboardButton(day, callback_data="ignore") for day in weekdays])
    
    # Получаем первый день месяца и количество дней
    first_day = datetime(year, month, 1)
    if month == 12:
        next_month = datetime(year + 1, 1, 1)
    else:
        next_month = datetime(year, month + 1, 1)
    days_in_month = (next_month - timedelta(days=1)).day
    
    # Определяем день недели для первого дня (1 = понедельник)
    first_weekday = first_day.isoweekday()
    
    # Добавляем пустые ячейки в начале
    row = []
    for _ in range(first_weekday - 1):
        row.append(InlineKeyboardButton(" ", callback_data="ignore"))
    
    # Добавляем дни месяца
    now = datetime.now()
    for day in range(1, days_in_month + 1):
        date_str = f"{year}-{month:02d}-{day:02d}"
        callback_data = f"select_custom_date_any_{date_str}"

        if day == now.day and month == now.month and year == now.year:
            row.append(InlineKeyboardButton(f"•{day}•", callback_data=callback_data))
        else:
            row.append(InlineKeyboardButton(str(day), callback_data=callback_data))

        if len(row) == 7:
            keyboard.row(*row)
            row = []

    # ДОБАВЛЯЕМ ОСТАВШИЕСЯ ТОЛЬКО ПОСЛЕ ЦИКЛА
    if row:
        while len(row) < 7:
            row.append(InlineKeyboardButton(" ", callback_data="ignore"))
        keyboard.row(*row)

    # Кнопки навигации
    nav_row = [
        InlineKeyboardButton("◀️", callback_data=f"prev_month_custom_any"),
        InlineKeyboardButton("🔙 Назад", callback_data="main_menu"),
        InlineKeyboardButton("▶️", callback_data=f"next_month_custom_any")
    ]
    keyboard.row(*nav_row)
    
    # Показываем инструкцию в зависимости от того, выбрана ли первая дата
    async with state.proxy() as data:
        first_date = data.get("first_date")
        if first_date is None:
            instruction = "Выберите две даты периода"
        else:
            instruction = f"Выбрана первая дата: {first_date.strftime('%d.%m.%Y')}\nВыберите вторую дату"
    
    try:
        # Пробуем отредактировать существующее сообщение
        await callback.message.edit_text(
            f"📅 <b>Выберите период</b>\n\n{instruction}:",
            reply_markup=keyboard
        )
    except MessageNotModified:
        await callback.answer()
    except Exception as e:
        print(f"Error in show_calendar_for_month_custom: {e}")
        # Если не удалось отредактировать, создаем новое сообщение
        try:
            await callback.message.answer(
                f"📅 <b>Выберите период</b>\n\n{instruction}:",
                reply_markup=keyboard
            )
        except Exception as e2:
            print(f"Error creating new message: {e2}")
            await callback.answer()

async def select_custom_date_callback(callback: types.CallbackQuery, state: FSMContext):
    """Обработчик выбора даты для кастомного периода"""
    parts = callback.data.split("_")
    date_str = parts[4]  # YYYY-MM-DD
    
    selected_date = datetime.strptime(date_str, "%Y-%m-%d")
    
    # Получаем текущие данные состояния
    data = await state.get_data()
    first_date = data.get("first_date")
    
    if first_date is None:
        # Первый клик - сохраняем первую дату
        await state.update_data(first_date=selected_date)
        await callback.answer(f"Выбрана первая дата: {selected_date.strftime('%d.%m.%Y')}")
    else:
        # Второй клик - сохраняем вторую дату и определяем период
        await state.update_data(second_date=selected_date)
        # Сразу обрабатываем выбор периода
        await process_custom_period_selection(callback, state)

async def process_custom_period_selection(callback: types.CallbackQuery, state: FSMContext):
    """Обрабатывает выбор двух дат и определяет период"""
    # Получаем данные состояния
    data = await state.get_data()
    first_date = data.get("first_date")
    second_date = data.get("second_date")
    
    if not first_date or not second_date:
        # Если одна из дат не выбрана, просто возвращаемся
        return
    
    # Если выбраны одинаковые даты, используем один день
    if first_date == second_date:
        start_date = first_date
        end_date = first_date
    else:
        # Определяем какая дата раньше, какая позже
        if first_date <= second_date:
            start_date = first_date
            end_date = second_date
        else:
            start_date = second_date
            end_date = first_date
    
    # Сохраняем период в state
    await state.update_data(
        custom_start_date=start_date,
        custom_end_date=end_date,
        custom_period=True
    )
    
    # Сразу подтверждаем выбор без промежуточного меню
    await confirm_custom_period_callback(callback, state)

async def show_custom_period_confirmation(callback: types.CallbackQuery, state: FSMContext, start_date, end_date):
    """Показывает подтверждение выбранного кастомного периода"""
    period_display = f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
    
    keyboard = InlineKeyboardMarkup(row_width=1)
    keyboard.add(
        InlineKeyboardButton("✅ Подтвердить", callback_data="confirm_custom_period"),
        InlineKeyboardButton("🔄 Выбрать заново", callback_data="restart_custom_period"),
        InlineKeyboardButton("🔙 Назад", callback_data="custom_period_back")
    )
    
    try:
        await callback.message.edit_text(
            f"📅 <b>Подтверждение выбранного периода</b>\n\n"
            f"Период: <b>{period_display}</b>\n\n"
            f"Нажмите 'Подтвердить' для расчета:",
            reply_markup=keyboard
        )
    except MessageNotModified:
        await callback.answer()
    except Exception as e:
        print(f"Error in show_custom_period_confirmation: {e}")
        await callback.answer()

async def confirm_custom_period_callback(callback: types.CallbackQuery, state: FSMContext):
    """Обработчик подтверждения кастомного периода"""
    # Получаем данные состояния
    data = await state.get_data()
    start_date = data["custom_start_date"]
    end_date = data["custom_end_date"]
    an_type = data.get("an_type", "an_1")  # Используем значение по умолчанию
    
    # Проверяем и исправляем an_type если нужно
    if not an_type or not an_type.startswith("an_"):
        an_type = "an_1"
        await state.update_data(an_type=an_type)
    
    # Создаем фейковый callback для select_anal_period_callback
    try:
        type_num = an_type.split('_')[1]
        callback.data = f"anperiod_custom_custom_{type_num}"
    except (IndexError, ValueError):
        callback.data = "anperiod_custom_custom_1"
    
    # Сохраняем кастомные даты в состоянии
    await state.update_data(
        custom_period=True,
        custom_start_date=start_date,
        custom_end_date=end_date,
        an_type=an_type
    )

    # Получаем данные магазина из состояния
    shop_data = data.get("shop")
    if not shop_data:
        await callback.answer("❌ Магазин не выбран. Сначала выберите магазин.", show_alert=True)
        return
        
    shop_id = shop_data["id"]
    api_token = shop_data["api_token"]
    
    # Вызываем основную функцию расчета
    await select_anal_period_callback(callback, state)

async def restart_custom_period_callback(callback: types.CallbackQuery, state: FSMContext):
    """Обработчик для выбора кастомного периода заново"""
    async with state.proxy() as data:
        data["date_selection_step"] = "first"
        # Очищаем предыдущие даты
        data.pop("first_date", None)
        data.pop("second_date", None)
        data.pop("custom_start_date", None)
        data.pop("custom_end_date", None)
    
    await show_calendar_for_custom_period(callback, state)

async def custom_period_navigation_callback(callback: types.CallbackQuery, state: FSMContext):
    """Обработчик навигации в календаре для кастомного периода"""
    parts = callback.data.split("_")
    action = parts[0]  # prev или next
    
    async with state.proxy() as data:
        current_month = data.get("calendar_month", datetime.now().month)
        current_year = data.get("calendar_year", datetime.now().year)
        
        if action == "prev":
            if current_month == 1:
                current_month = 12
                current_year -= 1
            else:
                current_month -= 1
        else:  # next
            if current_month == 12:
                current_month = 1
                current_year += 1
            else:
                current_month += 1
        
        data["calendar_month"] = current_month
        data["calendar_year"] = current_year
    
    try:
        await show_calendar_for_month_custom(callback, state, current_month, current_year, "any")
    except MessageNotModified:
        await callback.answer()
    except Exception as e:
        print(f"Error in custom_period_navigation_callback: {e}")
        await callback.answer()
